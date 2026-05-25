"""
build_tem_dcf.py  —  Generates TEM_dcf.xlsx matching the BBAI DCF template exactly.
Run: python build_tem_dcf.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── helpers ──────────────────────────────────────────────────────────────────
def s(ws, row, col, val, bold=False, italic=False, fmt=None, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=val)
    if bold or italic:
        c.font = Font(bold=bold, italic=italic)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if align:
        c.alignment = Alignment(wrap_text=(align=="wrap"), horizontal="left")
    return c

PCT  = '0.0%'
PCT1 = '0.00%'
INT  = '#,##0'
INT0 = '#,##0'
DOL  = '$#,##0.00'

BLUE   = "BDD7EE"   # input cell
YELLOW = "FFFF99"   # key assumption
GREY   = "D9D9D9"   # section header

def section(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor=GREY)

def hdr(ws, row, cols_vals, bold=True):
    for col, val in cols_vals:
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold)

# ── sensitivity formula builder ───────────────────────────────────────────────
def sens_formula(wacc, tgr, ufcf_rows="B122:B141", tv_row="B141",
                 debt_liq="Assumptions!C32,Assumptions!C27",
                 shares="Assumptions!C34"):
    w, g = wacc, tgr
    # sum of PV of UFCFs at mid-year convention
    terms = "+".join(
        f"B{122+i}/(1+{w})^{i+0.5}" for i in range(20)
    )
    tv    = f"{tv_row}*(1+{g})/({w}-{g})/(1+{w})^20"
    nd    = f"MAX(0,Assumptions!C32-Assumptions!C27)"
    nc    = f"MAX(0,Assumptions!C27-Assumptions!C32)"
    sh    = "Assumptions!C34"
    return f'=IF({w}>{g},({terms}+{tv}-{nd}+{nc})*1000/{sh},"-")'

# ── TEM data constants (all $thousands unless noted) ─────────────────────────

# Historical income statement
REV  = {23: 531822, 24: 693398, 25: 1271789, 22: 320668}
COR  = {23: 340802, 24: 479804, 25: 619999}   # derived: CostsAndExp - SGA - RD
SGA  = {23: 296760, 24: 755351, 25: 731738}
RD   = {23: 90343,  24: 149325, 25: 172924}
EBIT = {23:-196083, 24:-691082, 25:-252872}
DA   = {23: 21279,  24: 26356,  25: 32054}
SBC  = {23: 0,      24: 534138, 25: 124747}
CAPEX= {23: 34608,  24: 22121,  25: 21049}
CAPSW= {23: 6137,   24: 4292,   25: 5535}
INTEXP={23: 46869,  24: 53653,  25: 70267}
INTINC={23: 7601,   24: 11084,  25: 12628}
NETLOSS={23:-214118,24:-705809, 25:-245028}

# Balance sheet
CASH = {24: 340954,  25: 604787}
AR   = {24: 154819,  25: 311170}
GW   = {24: 73343,   25: 470211}
INTG = {24: 11716,   25: 355253}
DEBT = {24: 168192,  25: 208672}   # convertible notes
SHRS = {24: 119849000, 25: 174264000}  # raw share count

NOL  = 727301   # deferred tax assets $K ≈ federal NOL proxy

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 1: Assumptions
# ═══════════════════════════════════════════════════════════════════════════════
wa = wb.active
wa.title = "Assumptions"
wa.column_dimensions["A"].width = 52
wa.column_dimensions["B"].width = 14
wa.column_dimensions["C"].width = 14
wa.column_dimensions["D"].width = 14
wa.column_dimensions["E"].width = 48

# Title
s(wa, 1, 1, "TEM — DCF Model Assumptions", bold=True)
s(wa, 2, 1, "All dollar inputs in $thousands  |  Blue = input  |  Black = formula  |  Yellow = key assumption")

# ── A. HISTORICAL ─────────────────────────────────────────────────────────────
s(wa, 4, 1, "A.  HISTORICAL FINANCIAL DATA  (Source: TEM 10-K FY2025, SEC EDGAR CIK 0001717115)", bold=True)
hdr(wa, 5, [(1,"Line Item"),(2,"FY2023"),(3,"FY2024"),(4,"FY2025"),(5,"Notes")])

rows_hist = [
    (6,  "Revenue",            REV[23], REV[24], REV[25],  "Source: EDGAR RevenueFromContractWithCustomer"),
    (7,  "Cost of Revenue",    COR[23], COR[24], COR[25],  "Derived: CostsAndExpenses − SG&A − R&D"),
    (8,  "Gross Profit",       "=B6-B7","=C6-C7","=D6-D7","Formula: Revenue − CoR"),
    (9,  "Gross Margin %",     "=B8/B6","=C8/C6","=D8/D6","Formula"),
    (10, "SG&A",               SGA[23], SGA[24], SGA[25],  "Source: EDGAR GeneralAndAdministrativeExpense (incl SBC)"),
    (11, "R&D",                RD[23],  RD[24],  RD[25],   "Source: EDGAR ResearchAndDevelopmentExpense"),
    (12, "EBIT (Operating Loss)", EBIT[23], EBIT[24], EBIT[25], "Source: EDGAR OperatingIncomeLoss"),
    (13, "D&A",                DA[23],  DA[24],  DA[25],   "Source: EDGAR DepreciationDepletionAndAmortization"),
    (14, "Stock-Based Comp",   SBC[23], SBC[24], SBC[25],  "Source: EDGAR AllocatedShareBasedCompensationExpense"),
    (15, "CapEx (PP&E)",       CAPEX[23],CAPEX[24],CAPEX[25],"Source: EDGAR PaymentsToAcquirePropertyPlantAndEquipment"),
    (16, "Capitalized Software",CAPSW[23],CAPSW[24],CAPSW[25],"Source: EDGAR CapitalizedComputerSoftwareAdditions"),
    (17, "Interest Expense",   INTEXP[23],INTEXP[24],INTEXP[25],"Source: EDGAR InterestExpenseDebt"),
    (18, "Interest Income",    INTINC[23],INTINC[24],INTINC[25],"Source: EDGAR InvestmentIncomeInterest"),
    (19, "Net Loss",           NETLOSS[23],NETLOSS[24],NETLOSS[25],"Source: EDGAR NetIncomeLoss"),
]
for r, label, b, c, d, note in rows_hist:
    s(wa, r, 1, label)
    for col, val in [(2,b),(3,c),(4,d)]:
        cell = wa.cell(row=r, column=col, value=val)
        if r == 9:
            cell.number_format = PCT
        elif r not in (8,):
            cell.number_format = INT
    s(wa, r, 5, note)
# Gross Profit formatting
for col in [2,3,4]:
    wa.cell(row=8, column=col).number_format = INT

# ── A2. BALANCE SHEET ─────────────────────────────────────────────────────────
s(wa, 22, 1, "A2.  BALANCE SHEET SNAPSHOT  (Source: TEM 10-K FY2025, SEC EDGAR)", bold=True)
hdr(wa, 23, [(1,"Line Item"),(2,"FY2024"),(3,"FY2025"),(5,"Notes")])
bs_rows = [
    (24,"Cash & Equivalents",       CASH[24],  CASH[25],  "EDGAR CashAndCashEquivalentsAtCarryingValue"),
    (25,"AFS Investments (Current)",0,          0,         "TEM: none reported"),
    (26,"AFS Investments (Non-Current)",0,      0,         "TEM: none reported"),
    (27,"Total Liquidity",          "=B24+B25+B26","=C24+C25+C26","Formula: Cash + AFS"),
    (28,"Accounts Receivable",      AR[24],    AR[25],    "EDGAR AccountsReceivableNetCurrent"),
    (29,"Goodwill",                 GW[24],    GW[25],    "EDGAR Goodwill (inc. Ambry Genetics acq. FY2025)"),
    (30,"Intangibles, net",         INTG[24],  INTG[25],  "EDGAR IntangibleAssetsNetExcludingGoodwill"),
    (31,"Total Debt (Convertible Notes)",DEBT[24],DEBT[25],"EDGAR ConvertibleDebtNoncurrent"),
    (32,"Debt Used in Bridge",      "=B31",    "=C31",    "Convertible notes; no near-term forced conversion"),
    (33,"Shares Outstanding",       SHRS[24],  SHRS[25],  "EDGAR WeightedAverageDilutedShares × 1000"),
    (34,"Shares Used in Bridge",    "=B33",    "=C33",    "Formula: use FY2025 diluted count"),
    (35,"Federal NOL (DTA proxy, $K)",None,    NOL,       "EDGAR DeferredTaxAssetsGross FY2025"),
]
for r, label, b, c, note in bs_rows:
    s(wa, r, 1, label)
    for col, val in [(2,b),(3,c)]:
        if val is not None:
            cell = wa.cell(row=r, column=col, value=val)
            if r not in (32,34):
                cell.number_format = INT
    s(wa, r, 5, note)

# ── B. REVENUE GROWTH ─────────────────────────────────────────────────────────
s(wa, 38, 1, "B.  REVENUE GROWTH ASSUMPTIONS  (Yellow = key input)", bold=True)
hdr(wa, 39, [(1,"Assumption"),(2,"Base Case"),(3,"Bull Case"),(4,"Bear Case"),(5,"Notes")])
rev_growth = [
    (40,"FY2026 Revenue Growth Rate",0.35,0.45,0.20,"Moderation from 83% FY2025; Ambry now integrated"),
    (41,"FY2027 Revenue Growth Rate",0.30,0.38,0.15,"Data platform + genomics ramp"),
    (42,"FY2028 Revenue Growth Rate",0.25,0.32,0.12,"Continued AI/oncology expansion"),
    (43,"FY2029 Revenue Growth Rate",0.22,0.28,0.10,"Scale + payer contract wins"),
    (44,"FY2030 Revenue Growth Rate",0.20,0.25,0.08,"TAM penetration; platform maturation"),
    (45,"FY2031–2035 Revenue CAGR",  0.17,0.22,0.08,"Maturing growth; international entry"),
    (46,"FY2036–2040 Revenue CAGR",  0.11,0.15,0.06,"Large base; moderate compounding"),
    (47,"FY2041–2045 Revenue CAGR",  0.07,0.10,0.04,"Terminal approach; long-run normalization"),
]
for r, label, base, bull, bear, note in rev_growth:
    s(wa, r, 1, label)
    for col, val in [(2,base),(3,bull),(4,bear)]:
        c = wa.cell(row=r, column=col, value=val)
        c.number_format = PCT
        c.fill = PatternFill("solid", fgColor=YELLOW)
    s(wa, r, 5, note)

s(wa, 49, 1, "ACTIVE SCENARIO → Enter 1 = Base Case  |  2 = Bull Case  |  3 = Bear Case")
s(wa, 50, 1, "Active Scenario (1/2/3)")
c50 = wa.cell(row=50, column=2, value=1)
c50.fill = PatternFill("solid", fgColor=YELLOW)
c50.font = Font(bold=True)
s(wa, 50, 5, "Change this cell to switch scenarios on DCF sheet. 1=Base, 2=Bull, 3=Bear")

# ── C. MARGINS ────────────────────────────────────────────────────────────────
s(wa, 52, 1, "C.  MARGIN & COST ASSUMPTIONS", bold=True)
hdr(wa, 53, [(1,"Assumption"),(2,"Base Case"),(3,"Bull Case"),(4,"Bear Case"),(5,"Notes")])
margins = [
    (54,"Gross Margin % — FY2026",    0.53,0.57,0.48,"FY2025=51.2%; Ambry + AI platform scale"),
    (55,"Gross Margin % — FY2027",    0.56,0.60,0.50,"Platform mix shift; sequencing cost decline"),
    (56,"Gross Margin % — FY2028",    0.59,0.63,0.52,"Software-driven revenue grows faster"),
    (57,"Gross Margin % — FY2029",    0.62,0.66,0.54,"Scale economics"),
    (58,"Gross Margin % — FY2030",    0.64,0.69,0.56,"Mature platform mix"),
    (59,"Gross Margin % — FY2031–35", 0.67,0.72,0.58,"Long-run target; health-AI SaaS peers 65–75%"),
    (60,"Gross Margin % — FY2036–45", 0.70,0.76,0.60,"Terminal gross margin"),
    (61,"SG&A as % Revenue — FY2026", 0.48,0.42,0.56,"FY2025=57.5% (incl SBC); heavy sales investment"),
    (62,"SG&A as % Revenue — FY2027", 0.42,0.36,0.52,"Operating leverage begins"),
    (63,"SG&A as % Revenue — FY2028", 0.37,0.31,0.47,""),
    (64,"SG&A as % Revenue — FY2029", 0.33,0.27,0.43,""),
    (65,"SG&A as % Revenue — FY2030", 0.28,0.22,0.40,""),
    (66,"SG&A as % Revenue — FY2031–35",0.22,0.17,0.34,"Maturing cost structure"),
    (67,"SG&A as % Revenue — FY2036–45",0.18,0.13,0.28,"Terminal SG&A"),
    (68,"R&D as % Revenue — All Years",0.11,0.13,0.10,"FY2025=13.6%; normalizing"),
    (69,"D&A as % Revenue — All Years",0.08,0.07,0.09,"Elevated post-Ambry intangible amort (~$355M)"),
    (70,"CapEx as % Revenue — All Years",0.04,0.03,0.05,"Asset-light; FY2025 CapEx+CapSW=2.1% rev"),
    (71,"Stock-Based Comp as % Revenue",0.10,0.09,0.12,"FY2025=9.8%; normalizing from IPO spike"),
    (72,"Change in NWC as % Rev Change",0.06,0.04,0.09,"AR-heavy model; genomics payer lag"),
    (73,"Tax Rate (Cash)",              0.00,0.00,0.02,"$727M DTA / large NOL; ~0% cash tax 8–10 yrs"),
    (74,"Tax Rate (Terminal / WACC)",   0.21,0.21,0.21,"US statutory rate"),
]
for r, label, base, bull, bear, note in margins:
    s(wa, r, 1, label)
    for col, val in [(2,base),(3,bull),(4,bear)]:
        c = wa.cell(row=r, column=col, value=val)
        c.number_format = PCT
        if r in range(54,68):
            c.fill = PatternFill("solid", fgColor=BLUE)
        else:
            c.fill = PatternFill("solid", fgColor=BLUE)
    s(wa, r, 5, note)

# ── D. WACC ───────────────────────────────────────────────────────────────────
s(wa, 77, 1, "D.  WACC ASSUMPTIONS", bold=True)
hdr(wa, 78, [(1,"Assumption"),(2,"Base Case"),(3,"Bull Case"),(4,"Bear Case"),(5,"Notes")])
wacc_rows = [
    (79,"Risk-Free Rate",      0.043, 0.040, 0.046,"10-yr US Treasury; as of model date May 2026"),
    (80,"Equity Risk Premium", 0.055, 0.050, 0.060,"Damodaran US ERP estimate"),
    (81,"Beta (Levered)",      1.70,  1.40,  2.10, "Health-AI growth co.; post-IPO volatility"),
    (82,"Size / Illiquidity Premium",0.020,0.010,0.030,"Mid-cap premium; TEM mkt cap ~$8–12B range"),
    (83,"Cost of Equity",      "=B79+B81*B80+B82","=C79+C81*C80+C82","=D79+D81*D80+D82","Formula: Rf + Beta×ERP + Size"),
    (84,"Pre-Tax Cost of Debt",0.055, 0.050, 0.065,"Convertible notes; blended coupon est."),
    (85,"Weight of Equity",    0.85,  0.90,  0.80, "Post-debt capital structure"),
    (86,"Weight of Debt",      "=1-B85","=1-C85","=1-D85","Formula: 1 − Weight of Equity"),
    (87,"WACC",                "=B83*B85+B84*(1-D74)*B86","=C83*C85+C84*(1-D74)*C86","=D83*D85+D84*(1-D74)*D86","Formula: Ke×We + Kd×(1−t)×Wd"),
]
for r, label, base, bull, bear, note in wacc_rows:
    s(wa, r, 1, label)
    for col, val in [(2,base),(3,bull),(4,bear)]:
        c = wa.cell(row=r, column=col, value=val)
        if r in (79,80,82,84,85):
            c.number_format = PCT
            c.fill = PatternFill("solid", fgColor=BLUE)
        elif r == 81:
            c.number_format = '0.00'
            c.fill = PatternFill("solid", fgColor=BLUE)
        elif r in (83,86,87):
            c.number_format = PCT
    s(wa, r, 5, note)

# ── E. TERMINAL VALUE ─────────────────────────────────────────────────────────
s(wa, 90, 1, "E.  TERMINAL VALUE ASSUMPTIONS", bold=True)
hdr(wa, 91, [(1,"Assumption"),(2,"Base Case"),(3,"Bull Case"),(4,"Bear Case"),(5,"Notes")])
tv_rows = [
    (92,"Terminal Growth Rate (Gordon Growth)",0.035,0.040,0.025,"Must be < WACC; healthcare AI long-run GDP+ growth"),
    (93,"Terminal EV/EBITDA Multiple (cross-check)",20,28,12,"Health-tech/AI software comps 15–30×"),
    (94,"Terminal EBITDA Margin (for EV/EBITDA check)",0.28,0.35,0.18,"Long-run Adj. EBITDA margin target"),
]
for r, label, base, bull, bear, note in tv_rows:
    s(wa, r, 1, label)
    for col, val in [(2,base),(3,bull),(4,bear)]:
        c = wa.cell(row=r, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=YELLOW)
        if r in (92,94):
            c.number_format = PCT
    s(wa, r, 5, note)

# ── F. WACC DECOMPOSITION ─────────────────────────────────────────────────────
s(wa, 97, 1, "F.  WACC DECOMPOSITION — Why WACC is high & the two levers to lower it", bold=True)
hdr(wa, 98, [(1,"Component"),(2,"Current Value"),(3,"Source / Formula"),(4,"Contribution to WACC"),(5,"How to Lower It")])
wacc_decomp = [
    (99, "Risk-Free Rate (Rf)",    "=B79","Rf × We",               "=B79*B85","Falls automatically if rates cut"),
    (100,"Equity Risk Premium (ERP)","=B80","Damodaran ERP",       "=B81*B80*B85","Market-wide; not controllable"),
    (101,"Beta  ← #1 LEVER",       "=B81","Currently 1.70",        "=B81*B80*B85","Cut to 1.0–1.2 once FCF-positive, revenue diversified"),
    (102,"Size/Illiquidity ← #2",  "=B82","Currently 2%",          "=B82*B85","Drops to 0% once mkt cap >$5B consistently"),
    (103,"Cost of Equity (Ke)",    "=B83","Rf + Beta×ERP + Size",  "=B83*B85","Currently ~16%; driven by Beta and Size"),
    (104,"Cost of Debt (Kd after-tax)","=B84","Conv. notes ~5.5%", "=B84*(1-D74)*B86","Minor — 15% weight, convertible structure"),
    (105,"WACC  (Ke×We + Kd×(1−t)×Wd)","=B87","Formula result",  None,"PRIMARY: Beta 1.70→1.0 saves ~4pp; Size saves ~1.7pp"),
]
for r, comp, cur, src, contrib, how in wacc_decomp:
    s(wa, r, 1, comp)
    s(wa, r, 2, cur)
    s(wa, r, 3, src)
    if contrib:
        wa.cell(row=r, column=4, value=contrib).number_format = PCT
    s(wa, r, 5, how)

# ── G. WACC COMPRESSION ROADMAP ───────────────────────────────────────────────
s(wa, 107, 1, "G.  WACC COMPRESSION ROADMAP — Change Beta (row 81) and Size Premium (row 82) to model each stage", bold=True)
hdr(wa, 108, [(1,"Stage"),(2,"Beta"),(3,"Size Premium"),(4,"Implied WACC"),(5,"Trigger / Condition")])
stages = [
    (109,"Current Base Case",     "=B81","=B82","=B87",            "Pre-profit; heavy SBC; concentrated revenue"),
    (110,"Stage 1 — FCF Break-Even",1.30, 0.015,"=(B79+1.30*B80+0.015)*B85+B84*(1-D74)*B86","TEM FCF-positive ~FY2029E"),
    (111,"Stage 2 — Proven Platform",1.00,0.005,"=(B79+1.00*B80+0.005)*B85+B84*(1-D74)*B86","ARR visible; genomics recurring revenue scaling"),
    (112,"Stage 3 — Re-Rating ($50+)",0.80,0.000,"=(B79+0.80*B80+0.000)*B85+B84*(1-D74)*B86","Debt-free, profitable 3+ yrs, category leader"),
    (113,"Blue Sky — Epic/PLTR Comp",0.60,0.000,"=(B79+0.60*B80+0.000)*B85+B84*(1-D74)*B86","Tempus moat/brand = health-AI infrastructure"),
]
for r, stage, beta, size, wacc_f, trigger in stages:
    s(wa, r, 1, stage)
    wa.cell(row=r, column=2, value=beta).number_format = '0.00'
    wa.cell(row=r, column=3, value=size).number_format = PCT
    wa.cell(row=r, column=4, value=wacc_f).number_format = PCT
    s(wa, r, 5, trigger)

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 2: DCF Model
# ═══════════════════════════════════════════════════════════════════════════════
wd = wb.create_sheet("DCF Model")
wd.column_dimensions["A"].width = 46
for i in range(2, 23):
    wd.column_dimensions[get_column_letter(i)].width = 12

COLS = list("BCDEFGHIJKLMNOPQRSTUV")   # B=FY2025 actual, C–V = FY2026–FY2045
YEARS = ["FY2025\n(Actual)"] + [str(y) for y in range(2026, 2046)]

# Title rows
s(wd, 1, 1, "TEM — 20-Year Unlevered Free Cash Flow DCF Model", bold=True)
s(wd, 2, 1, '=IF(Assumptions!B50=1,"▶  ACTIVE: BASE CASE",IF(Assumptions!B50=2,"▶  ACTIVE: BULL CASE","▶  ACTIVE: BEAR CASE"))')

# Year headers (row 3)
s(wd, 3, 1, "Year", bold=True)
for i, yr in enumerate(YEARS):
    c = wd.cell(row=3, column=2+i, value=yr)
    c.font = Font(bold=True)
    c.alignment = Alignment(wrap_text=True)

# ── A. REVENUE ────────────────────────────────────────────────────────────────
s(wd, 4, 1, "A.  REVENUE PROJECTION", bold=True)
s(wd, 5, 1, "Revenue Growth Rate (Active Scenario)")
s(wd, 5, 2, "—")

# Growth rate CHOOSE by year group
asmp_grow = {  # DCF col index (0=FY2025 in B) → Assumptions row
    1: 40, 2: 41, 3: 42, 4: 43, 5: 44,
    6: 45, 7: 45, 8: 45, 9: 45, 10: 45,
    11: 46, 12: 46, 13: 46, 14: 46, 15: 46,
    16: 47, 17: 47, 18: 47, 19: 47, 20: 47,
}
for i in range(1, 21):   # i=0 is FY2025 col B
    col = i + 2   # C=3 … V=22
    ar = asmp_grow[i]
    wd.cell(row=5, column=col,
            value=f"=CHOOSE(Assumptions!B50,Assumptions!B{ar},Assumptions!C{ar},Assumptions!D{ar})"
           ).number_format = PCT

s(wd, 6, 1, "Revenue ($000s)")
wd.cell(row=6, column=2, value="=Assumptions!D6").number_format = INT
for i in range(1, 21):
    col = i + 2
    prev = get_column_letter(col - 1)
    wd.cell(row=6, column=col,
            value=f"={prev}6*(1+{get_column_letter(col)}5)").number_format = INT

# ── B. INCOME STATEMENT ───────────────────────────────────────────────────────
s(wd, 7, 1, "B.  INCOME STATEMENT PROJECTION", bold=True)

# Gross Margin %
s(wd, 8, 1, "   Gross Margin %")
wd.cell(row=8, column=2, value="=Assumptions!D9").number_format = PCT
gm_rows  = {1:54,2:55,3:56,4:57,5:58,6:59,7:59,8:59,9:59,10:59,
            11:60,12:60,13:60,14:60,15:60,16:60,17:60,18:60,19:60,20:60}
for i in range(1, 21):
    col = i + 2
    ar = gm_rows[i]
    wd.cell(row=8, column=col,
            value=f"=CHOOSE(Assumptions!B50,Assumptions!B{ar},Assumptions!C{ar},Assumptions!D{ar})"
           ).number_format = PCT

# Gross Profit
s(wd, 9, 1, "   Gross Profit ($000s)")
wd.cell(row=9, column=2, value="=Assumptions!D8").number_format = INT
for i in range(1, 21):
    col = i + 2
    lc = get_column_letter(col)
    wd.cell(row=9, column=col, value=f"={lc}6*{lc}8").number_format = INT

# SG&A %
s(wd, 10, 1, "   SG&A as % Revenue")
wd.cell(row=10, column=2, value="=Assumptions!D10/Assumptions!D6").number_format = PCT
sga_rows = {1:61,2:62,3:63,4:64,5:65,6:66,7:66,8:66,9:66,10:66,
            11:67,12:67,13:67,14:67,15:67,16:67,17:67,18:67,19:67,20:67}
for i in range(1, 21):
    col = i + 2
    ar = sga_rows[i]
    wd.cell(row=10, column=col,
            value=f"=CHOOSE(Assumptions!B50,Assumptions!B{ar},Assumptions!C{ar},Assumptions!D{ar})"
           ).number_format = PCT

# SG&A $
s(wd, 11, 1, "   SG&A ($000s)")
wd.cell(row=11, column=2, value="=Assumptions!D10").number_format = INT
for i in range(1, 21):
    col = i + 2; lc = get_column_letter(col)
    wd.cell(row=11, column=col, value=f"={lc}6*{lc}10").number_format = INT

# R&D %
s(wd, 12, 1, "   R&D as % Revenue")
wd.cell(row=12, column=2, value="=Assumptions!D11/Assumptions!D6").number_format = PCT
for i in range(1, 21):
    col = i + 2
    wd.cell(row=12, column=col,
            value="=CHOOSE(Assumptions!B50,Assumptions!B68,Assumptions!C68,Assumptions!D68)"
           ).number_format = PCT

# R&D $
s(wd, 13, 1, "   R&D ($000s)")
wd.cell(row=13, column=2, value="=Assumptions!D11").number_format = INT
for i in range(1, 21):
    col = i + 2; lc = get_column_letter(col)
    wd.cell(row=13, column=col, value=f"={lc}6*{lc}12").number_format = INT

# EBIT
s(wd, 14, 1, "EBIT ($000s)", bold=True)
wd.cell(row=14, column=2, value="=Assumptions!D12").number_format = INT
for i in range(1, 21):
    col = i + 2; lc = get_column_letter(col)
    wd.cell(row=14, column=col, value=f"={lc}9-{lc}11-{lc}13").number_format = INT

# EBIT Margin
s(wd, 15, 1, "   EBIT Margin %")
wd.cell(row=15, column=2, value="=Assumptions!D12/Assumptions!D6").number_format = PCT
for i in range(1, 21):
    col = i + 2; lc = get_column_letter(col)
    wd.cell(row=15, column=col, value=f"={lc}14/{lc}6").number_format = PCT

# Cash Tax Rate
s(wd, 16, 1, "   Cash Tax Rate (Active Scenario)")
wd.cell(row=16, column=2, value="=0").number_format = PCT
for i in range(1, 21):
    col = i + 2
    wd.cell(row=16, column=col,
            value="=CHOOSE(Assumptions!B50,Assumptions!B73,Assumptions!C73,Assumptions!D73)"
           ).number_format = PCT

# NOPAT
s(wd, 17, 1, "NOPAT = EBIT × (1 − Tax Rate) ($000s)", bold=True)
for i in range(0, 21):
    col = i + 2; lc = get_column_letter(col)
    wd.cell(row=17, column=col, value=f"={lc}14*(1-{lc}16)").number_format = INT

# ── C. FCF BUILD ──────────────────────────────────────────────────────────────
s(wd, 18, 1, "C.  UNLEVERED FREE CASH FLOW BUILD", bold=True)

# D&A %
s(wd, 19, 1, "   D&A as % Revenue")
wd.cell(row=19, column=2, value="=Assumptions!D13/Assumptions!D6").number_format = PCT
for i in range(1, 21):
    col = i + 2
    wd.cell(row=19, column=col,
            value="=CHOOSE(Assumptions!B50,Assumptions!B69,Assumptions!C69,Assumptions!D69)"
           ).number_format = PCT

# (+) D&A
s(wd, 20, 1, "   (+) D&A ($000s)")
wd.cell(row=20, column=2, value="=Assumptions!D13").number_format = INT
for i in range(1, 21):
    col = i + 2; lc = get_column_letter(col)
    wd.cell(row=20, column=col, value=f"={lc}6*{lc}19").number_format = INT

# CapEx %
s(wd, 21, 1, "   CapEx as % Revenue")
wd.cell(row=21, column=2, value="=(Assumptions!D15+Assumptions!D16)/Assumptions!D6").number_format = PCT
for i in range(1, 21):
    col = i + 2
    wd.cell(row=21, column=col,
            value="=CHOOSE(Assumptions!B50,Assumptions!B70,Assumptions!C70,Assumptions!D70)"
           ).number_format = PCT

# (-) CapEx
s(wd, 22, 1, "   (−) CapEx + Capitalized Software ($000s)")
wd.cell(row=22, column=2, value="=Assumptions!D15+Assumptions!D16").number_format = INT
for i in range(1, 21):
    col = i + 2; lc = get_column_letter(col)
    wd.cell(row=22, column=col, value=f"={lc}6*{lc}21").number_format = INT

# NWC %
s(wd, 23, 1, "   Δ NWC as % Revenue Change")
wd.cell(row=23, column=2, value="=0").number_format = PCT
for i in range(1, 21):
    col = i + 2
    wd.cell(row=23, column=col,
            value="=CHOOSE(Assumptions!B50,Assumptions!B72,Assumptions!C72,Assumptions!D72)"
           ).number_format = PCT

# (-) NWC
s(wd, 24, 1, "   (−) Change in NWC ($000s)")
wd.cell(row=24, column=2, value="=0").number_format = INT
for i in range(1, 21):
    col = i + 2; lc = get_column_letter(col)
    prev = get_column_letter(col - 1)
    wd.cell(row=24, column=col, value=f"=({lc}6-{prev}6)*{lc}23").number_format = INT

# UFCF
s(wd, 25, 1, "UNLEVERED FCF ($000s)", bold=True)
for i in range(0, 21):
    col = i + 2; lc = get_column_letter(col)
    wd.cell(row=25, column=col, value=f"={lc}17+{lc}20-{lc}22-{lc}24").number_format = INT

# UFCF Margin
s(wd, 26, 1, "   UFCF Margin %")
for i in range(0, 21):
    col = i + 2; lc = get_column_letter(col)
    wd.cell(row=26, column=col, value=f"=IF({lc}6<>0,{lc}25/{lc}6,0)").number_format = PCT

# ── D. DISCOUNTING ────────────────────────────────────────────────────────────
s(wd, 27, 1, "D.  DISCOUNTING  (Mid-Year Convention)", bold=True)

# WACC
s(wd, 28, 1, "WACC (Active Scenario)")
wd.cell(row=28, column=2,
        value="=CHOOSE(Assumptions!B50,Assumptions!B87,Assumptions!C87,Assumptions!D87)"
       ).number_format = PCT
for i in range(1, 21):
    col = i + 2
    wd.cell(row=28, column=col,
            value="=CHOOSE(Assumptions!B50,Assumptions!B87,Assumptions!C87,Assumptions!D87)"
           ).number_format = PCT

# Discount Period
s(wd, 29, 1, "   Discount Period (Mid-Year)")
wd.cell(row=29, column=2, value="—")
for i in range(1, 21):
    col = i + 2
    wd.cell(row=29, column=col, value=f"={i}-0.5").number_format = '0.0'

# Discount Factor
s(wd, 30, 1, "   Discount Factor")
wd.cell(row=30, column=2, value="—")
for i in range(1, 21):
    col = i + 2; lc = get_column_letter(col)
    wd.cell(row=30, column=col, value=f"=1/(1+{lc}28)^{lc}29").number_format = '0.0000'

# PV of UFCF
s(wd, 31, 1, "   PV of UFCF ($000s)")
wd.cell(row=31, column=2, value="—")
for i in range(1, 21):
    col = i + 2; lc = get_column_letter(col)
    wd.cell(row=31, column=col, value=f"={lc}25*{lc}30").number_format = INT

# Sum PV
s(wd, 32, 1, "Sum of PV (UFCF) — Years 1–20 ($000s)", bold=True)
wd.cell(row=32, column=2, value="=SUM(C31:V31)").number_format = INT

# ── E. TERMINAL VALUE ─────────────────────────────────────────────────────────
s(wd, 33, 1, "E.  TERMINAL VALUE  (Gordon Growth Model — Primary)", bold=True)
s(wd, 34, 1, "   Year 20 UFCF ($000s)"); wd.cell(row=34, column=2, value="=V25").number_format = INT
s(wd, 35, 1, "   Terminal Growth Rate (Active Scenario)")
wd.cell(row=35, column=2,
        value="=CHOOSE(Assumptions!B50,Assumptions!B92,Assumptions!C92,Assumptions!D92)"
       ).number_format = PCT
s(wd, 36, 1, "   WACC (Active Scenario)")
wd.cell(row=36, column=2,
        value="=CHOOSE(Assumptions!B50,Assumptions!B87,Assumptions!C87,Assumptions!D87)"
       ).number_format = PCT
s(wd, 37, 1, "Terminal Value (undiscounted) ($000s)", bold=True)
wd.cell(row=37, column=2, value="=B34*(1+B35)/(B36-B35)").number_format = INT
s(wd, 38, 1, "   Terminal Value Discount Factor (Year 20)")
wd.cell(row=38, column=2, value="=1/(1+V28)^20").number_format = '0.0000'
s(wd, 39, 1, "PV of Terminal Value ($000s)", bold=True)
wd.cell(row=39, column=2, value="=B37*B38").number_format = INT
s(wd, 40, 1, "   Terminal Value as % of Total Enterprise Value")
wd.cell(row=40, column=2, value="=B39/(B32+B39)").number_format = PCT

# EV/EBITDA cross-check
s(wd, 41, 1, "Terminal Value Cross-Check: EV/EBITDA Exit Multiple")
s(wd, 42, 1, "   Year 20 Revenue ($000s)"); wd.cell(row=42, column=2, value="=V6").number_format = INT
s(wd, 43, 1, "   Terminal EBITDA Margin (Active Scenario)")
wd.cell(row=43, column=2,
        value="=CHOOSE(Assumptions!B50,Assumptions!B94,Assumptions!C94,Assumptions!D94)"
       ).number_format = PCT
s(wd, 44, 1, "   Implied Year 20 EBITDA ($000s)"); wd.cell(row=44, column=2, value="=B42*B43").number_format = INT
s(wd, 45, 1, "   Exit EV/EBITDA Multiple (Active Scenario)")
wd.cell(row=45, column=2,
        value="=CHOOSE(Assumptions!B50,Assumptions!B93,Assumptions!C93,Assumptions!D93)"
       ).number_format = '0.0x'
s(wd, 46, 1, "   Terminal Value — EV/EBITDA Method (undiscounted) ($000s)")
wd.cell(row=46, column=2, value="=B44*B45").number_format = INT
s(wd, 47, 1, "   PV of Terminal Value — EV/EBITDA Method ($000s)")
wd.cell(row=47, column=2, value="=B46*B38").number_format = INT

# ── F. EQUITY BRIDGE ──────────────────────────────────────────────────────────
s(wd, 48, 1, "F.  ENTERPRISE VALUE  →  EQUITY VALUE  →  IMPLIED SHARE PRICE", bold=True)
s(wd, 49, 1, "Enterprise Value — Gordon Growth ($000s)", bold=True)
wd.cell(row=49, column=2, value="=B32+B39").number_format = INT
s(wd, 50, 1, "   Enterprise Value — EV/EBITDA Cross-Check ($000s)")
wd.cell(row=50, column=2, value="=B32+B47").number_format = INT
s(wd, 51, 1, "   (−) Net Debt (if Debt > Liquidity) ($000s)")
wd.cell(row=51, column=2, value="=MAX(0,Assumptions!C32-Assumptions!C27)").number_format = INT
wd.cell(row=51, column=3, value="If liquidity > debt, TEM is net cash — add back in equity bridge")
s(wd, 52, 1, "   (+) Net Cash (if Liquidity > Debt) ($000s)")
wd.cell(row=52, column=2, value="=MAX(0,Assumptions!C27-Assumptions!C32)").number_format = INT
s(wd, 53, 1, "Equity Value — Gordon Growth ($000s)", bold=True)
wd.cell(row=53, column=2, value="=B49-B51+B52").number_format = INT
s(wd, 54, 1, "   Diluted Shares Outstanding")
wd.cell(row=54, column=2, value="=Assumptions!C34").number_format = INT
s(wd, 55, 1, "IMPLIED SHARE PRICE ($)", bold=True)
wd.cell(row=55, column=2, value="=IF(B54>0,(B53*1000)/B54,0)").number_format = DOL

# ── G. SCENARIO SUMMARY ───────────────────────────────────────────────────────
s(wd, 57, 1, "G.  SCENARIO SUMMARY  (Run each scenario by changing Active Scenario on Assumptions sheet)", bold=True)
hdr(wd, 58, [(1,"Metric"),(2,"Base Case\n(Scenario 1)"),(3,"Bull Case\n(Scenario 2)"),(4,"Bear Case\n(Scenario 3)")])
for col in [2,3,4]:
    wd.cell(row=58, column=col).alignment = Alignment(wrap_text=True)

scen_rows = [
    (59,"FY2026E Revenue ($000s)",        "=B102","=C102","=D102"),
    (60,"FY2026E Revenue Growth %",       "=Assumptions!B40","=Assumptions!C40","=Assumptions!D40"),
    (61,"FY2030E Revenue ($000s)",        "=B106","=C106","=D106"),
    (62,"FY2045E Revenue ($000s)",        "=B121","=C121","=D121"),
    (63,"FY2026E Gross Margin %",         "=Assumptions!B54","=Assumptions!C54","=Assumptions!D54"),
    (64,"FY2026E EBIT ($000s)",           "=B102*(Assumptions!B54-Assumptions!B61-Assumptions!B68)",
                                          "=C102*(Assumptions!C54-Assumptions!C61-Assumptions!C68)",
                                          "=D102*(Assumptions!D54-Assumptions!D61-Assumptions!D68)"),
    (65,"WACC",                           "=Assumptions!B87","=Assumptions!C87","=Assumptions!D87"),
    (66,"Terminal Growth Rate",           "=Assumptions!B92","=Assumptions!C92","=Assumptions!D92"),
    (67,"Sum PV of FCFs ($000s)",         "=B162","=C162","=D162"),
    (68,"PV of Terminal Value ($000s)",   "=B164","=C164","=D164"),
    (69,"Enterprise Value ($000s)",       "=B165","=C165","=D165"),
    (70,"Equity Value ($000s)",           "=B166","=C166","=D166"),
    (71,"Implied Share Price ($)",        "=B167","=C167","=D167"),
]
pct_rows_g = {60,63,65,66}
dol_rows_g = {71}
for r, label, b, c, d in scen_rows:
    s(wd, r, 1, label)
    for col, val in [(2,b),(3,c),(4,d)]:
        cell = wd.cell(row=r, column=col, value=val)
        if r in pct_rows_g: cell.number_format = PCT
        elif r in dol_rows_g: cell.number_format = DOL
        else: cell.number_format = INT

# ── H. SENSITIVITY TABLE ──────────────────────────────────────────────────────
s(wd, 74, 1, "H.  SENSITIVITY: Implied Share Price ($)  —  WACC vs Terminal Growth Rate  (Base Case revenue/margins)", bold=True)
hdr(wd, 75, [(1,"WACC \\ TGR →"),(2,"2.0%"),(3,"2.5%"),(4,"3.0%"),(5,"3.5%"),(6,"4.0%"),(7,"4.5%")])
wacc_vals = [0.10, 0.12, 0.14, 0.16, 0.18, 0.20]
tgr_vals  = [0.020,0.025,0.030,0.035,0.040,0.045]
for ri, w in enumerate(wacc_vals):
    row = 76 + ri
    wd.cell(row=row, column=1, value=f"{int(w*100)}%").font = Font(bold=True)
    for ci, g in enumerate(tgr_vals):
        col = 2 + ci
        wd.cell(row=row, column=col,
                value=sens_formula(w, g)).number_format = DOL

# ── SCENARIO ENGINE (helper rows 100–167) ─────────────────────────────────────
s(wd, 100, 1, "── SCENARIO ENGINE (helper rows – do not edit) ──", bold=True)
hdr(wd, 101, [(1,"Metric"),(2,"Base"),(3,"Bull"),(4,"Bear")])

# Revenue rows 102-121 (FY2026–FY2045)
asmp_r = [40,41,42,43,44,45,45,45,45,45,46,46,46,46,46,47,47,47,47,47]
for i, ar in enumerate(asmp_r):
    row = 102 + i
    yr  = 2026 + i
    s(wd, row, 1, f"Rev {yr}")
    if i == 0:
        b = f"=Assumptions!D6*(1+Assumptions!B{ar})"
        c = f"=Assumptions!D6*(1+Assumptions!C{ar})"
        d = f"=Assumptions!D6*(1+Assumptions!D{ar})"
    else:
        prev = 101 + i
        b = f"=B{prev}*(1+Assumptions!B{ar})"
        c = f"=C{prev}*(1+Assumptions!C{ar})"
        d = f"=D{prev}*(1+Assumptions!D{ar})"
    for col, val in [(2,b),(3,c),(4,d)]:
        wd.cell(row=row, column=col, value=val).number_format = INT

# UFCF helper rows 122-141 (FY2026–FY2045)
# GM row map per year:
gm_asmp  = [54,55,56,57,58,59,59,59,59,59,60,60,60,60,60,60,60,60,60,60]
sga_asmp = [61,62,63,64,65,66,66,66,66,66,67,67,67,67,67,67,67,67,67,67]

for i in range(20):
    row = 122 + i
    yr  = 2026 + i
    ga  = gm_asmp[i]
    sa  = sga_asmp[i]
    rev_r   = 102 + i
    prev_r  = 101 + i  # previous revenue row (101 = Assumptions!D6 for i=0)
    s(wd, row, 1, f"UFCF {yr}")

    def ufcf_f(sc):
        """sc = 'B' (base), 'C' (bull), 'D' (bear)"""
        rev     = f"{sc}{rev_r}"
        gm      = f"Assumptions!{sc}{ga}"
        sg      = f"Assumptions!{sc}{sa}"
        rd      = f"Assumptions!{sc}68"
        tx      = f"Assumptions!{sc}73"
        da      = f"Assumptions!{sc}69"
        cx      = f"Assumptions!{sc}70"
        nwc     = f"Assumptions!{sc}72"
        prev_rev = f"Assumptions!D6" if i == 0 else f"{sc}{prev_r}"
        return (f"=(({rev}*{gm}-{rev}*{sg}-{rev}*{rd})*(1-{tx}))"
                f"+({rev}*{da})-({rev}*{cx})"
                f"-(({rev}-{prev_rev})*{nwc})")

    for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
        wd.cell(row=row, column=col, value=ufcf_f(sc)).number_format = INT

# PV of UFCF helper rows 142-161
for i in range(20):
    row = 142 + i
    yr  = 2026 + i
    ufcf_r = 122 + i
    t = i + 0.5    # mid-year period
    s(wd, row, 1, f"PV_UFCF {yr}")
    for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
        wacc_ref = f"Assumptions!{sc}87"
        wd.cell(row=row, column=col,
                value=f"={sc}{ufcf_r}/(1+{wacc_ref})^{t}"
               ).number_format = INT

# Summary rows 162-167
s(wd, 162, 1, "Sum PV_UFCF", bold=True)
for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
    wd.cell(row=162, column=col, value=f"=SUM({sc}142:{sc}161)").number_format = INT

s(wd, 163, 1, "")  # spacer

s(wd, 164, 1, "PV of Terminal Value", bold=True)
for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
    tv_ufcf = f"{sc}141"
    tgr_ref  = f"Assumptions!{sc}92"
    wacc_ref = f"Assumptions!{sc}87"
    wd.cell(row=164, column=col,
            value=f"={tv_ufcf}*(1+{tgr_ref})/({wacc_ref}-{tgr_ref})/(1+{wacc_ref})^20"
           ).number_format = INT

s(wd, 165, 1, "Enterprise Value", bold=True)
for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
    wd.cell(row=165, column=col, value=f"={sc}162+{sc}164").number_format = INT

s(wd, 166, 1, "Equity Value", bold=True)
nd = "MAX(0,Assumptions!C32-Assumptions!C27)"
nc = "MAX(0,Assumptions!C27-Assumptions!C32)"
for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
    wd.cell(row=166, column=col,
            value=f"={sc}165-{nd}+{nc}").number_format = INT

s(wd, 167, 1, "Implied Share Price ($)", bold=True)
for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
    wd.cell(row=167, column=col,
            value=f"=IF(Assumptions!C34>0,({sc}166*1000)/Assumptions!C34,0)"
           ).number_format = DOL

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save("/home/user/Modelingagent1/TEM_dcf.xlsx")
print("Saved: TEM_dcf.xlsx")

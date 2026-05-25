#!/usr/bin/env python3
"""
build_dcf.py  —  Generic DCF workbook builder (BBAI template structure).

Called by watchdog.py for each ticker.  Also importable as a library.

Usage:
  python build_dcf.py --ticker TEM
  python build_dcf.py --ticker RGTI
  python build_dcf.py --ticker BBAI
"""

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Style helpers ─────────────────────────────────────────────────────────────

PCT   = '0.0%'
INT   = '#,##0'
DOL   = '$#,##0.00'
BLUE  = "BDD7EE"
YELLOW= "FFFF99"
GREY  = "D9D9D9"

def _c(ws, row, col, val, bold=False, fmt=None, fill=None, wrap=False):
    cell = ws.cell(row=row, column=col, value=val)
    if bold:            cell.font = Font(bold=True)
    if fmt:             cell.number_format = fmt
    if fill:            cell.fill = PatternFill("solid", fgColor=fill)
    if wrap:            cell.alignment = Alignment(wrap_text=True)
    return cell

def _hdr(ws, row, cols_vals):
    for col, val in cols_vals:
        ws.cell(row=row, column=col, value=val).font = Font(bold=True)

def _section(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(bold=True)
    c.fill  = PatternFill("solid", fgColor=GREY)

def _input(ws, row, col, val, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=BLUE)
    if fmt: c.number_format = fmt
    return c

def _key(ws, row, col, val, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=YELLOW)
    if fmt: c.number_format = fmt
    return c

# ── Sensitivity formula ───────────────────────────────────────────────────────

def _sens(w, g):
    """Implied share price formula for a given hardcoded WACC w and TGR g."""
    terms = "+".join(f"B{122+i}/(1+{w})^{i+0.5}" for i in range(20))
    tv    = f"B141*(1+{g})/({w}-{g})/(1+{w})^20"
    nd    = "MAX(0,Assumptions!C32-Assumptions!C27)"
    nc    = "MAX(0,Assumptions!C27-Assumptions!C32)"
    sh    = "Assumptions!C34"
    return f'=IF({w}>{g},({terms}+{tv}-{nd}+{nc})*1000/{sh},"-")'

# ── Core builder ─────────────────────────────────────────────────────────────

def build_workbook(cfg: dict) -> openpyxl.Workbook:
    """
    Build a DCF workbook from *cfg*.

    cfg keys
    --------
    ticker      str   e.g. "TEM"
    cik         str   e.g. "0001717115"
    source_note str   e.g. "TEM 10-K FY2025, SEC EDGAR"

    # Historical income statement  ($thousands)
    rev    dict  {23: val, 24: val, 25: val}   (keyed by 2-digit FY)
    cor    dict  {23: val, 24: val, 25: val}
    sga    dict  ...
    rd     dict  ...
    ebit   dict  ...
    da     dict  ...
    sbc    dict  ...
    capex  dict  ...
    capsw  dict  ...
    intexp dict  ...
    intinc dict  ...
    netloss dict ...
    adj_ebitda dict  optional, {24: val, 25: val}

    # Balance sheet  ($thousands)
    cash_24 / cash_25
    afs_curr_24 / afs_curr_25     AFS investments current
    afs_nc_24   / afs_nc_25       AFS investments non-current
    ar_24 / ar_25
    goodwill_24 / goodwill_25
    intg_24 / intg_25
    debt_24 / debt_25             gross principal
    post_debt_25                  debt used in equity bridge (may differ)
    shares_24 / shares_25         raw share count
    post_shares_25                shares used in equity bridge
    nol                           NOL carryforward ($K)
    post_debt_note                text note for row 32
    post_shares_note              text note for row 34

    # Revenue growth  (Base / Bull / Bear)
    grow = {
      26: (base, bull, bear),  27: ..., 28: ..., 29: ..., 30: ...,
      '3135': (...), '3640': (...), '4145': (...)
    }
    grow_notes = {26: "...", ...}  optional

    # Gross margin %
    gm = {26: (b,u,r), ..., '3135':..., '3640-45':...}
    gm_notes = {...}

    # SG&A %
    sga_pct = {26:..., ..., '3135':..., '3640-45':...}
    sga_notes = {...}

    # Cost ratios  (single Base/Bull/Bear tuple)
    rd_pct      (b, u, r)
    da_pct      (b, u, r)
    capex_pct   (b, u, r)
    sbc_pct     (b, u, r)
    nwc_pct     (b, u, r)
    cash_tax    (b, u, r)
    notes_cost  dict  optional

    # WACC
    rf          (b, u, r)
    erp         (b, u, r)
    beta        (b, u, r)
    size_prem   (b, u, r)
    kd          (b, u, r)
    we          (b, u, r)
    wacc_notes  dict  optional

    # Terminal value
    tgr         (b, u, r)
    ev_ebitda   (b, u, r)
    ebitda_marg (b, u, r)

    # WACC decomp narrative
    wacc_decomp_rows  list of (label, how_to_lower)

    # WACC compression roadmap
    wacc_stages  list of (label, beta, size, trigger)
    """
    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════════
    #  SHEET 1 — Assumptions
    # ════════════════════════════════════════════════════════════════════════
    wa = wb.active
    wa.title = "Assumptions"
    wa.column_dimensions["A"].width = 54
    wa.column_dimensions["B"].width = 14
    wa.column_dimensions["C"].width = 14
    wa.column_dimensions["D"].width = 14
    wa.column_dimensions["E"].width = 50

    tk  = cfg["ticker"]
    src = cfg.get("source_note", f"{tk} 10-K FY2025, SEC EDGAR")

    _c(wa, 1, 1, f"{tk} — DCF Model Assumptions", bold=True)
    _c(wa, 2, 1, "All dollar inputs in $thousands  |  Blue = input  |  Black = formula  |  Yellow = key assumption")

    # ── A. Historical ────────────────────────────────────────────────────────
    _c(wa, 4, 1, f"A.  HISTORICAL FINANCIAL DATA  (Source: {src})", bold=True)
    _hdr(wa, 5, [(1,"Line Item"),(2,"FY2023"),(3,"FY2024"),(4,"FY2025"),(5,"Notes")])

    def _hist_row(row, label, b, c, d, note="", pct=False):
        _c(wa, row, 1, label)
        fmt = PCT if pct else INT
        for col, val in [(2,b),(3,c),(4,d)]:
            wa.cell(row=row, column=col, value=val).number_format = fmt
        _c(wa, row, 5, note)

    rev  = cfg["rev"];  cor = cfg["cor"];  sga_h = cfg["sga"]
    rd_h = cfg["rd"];   ebt = cfg["ebit"]; da_h  = cfg["da"]
    sbc_h= cfg["sbc"];  cx_h= cfg["capex"];csw_h = cfg["capsw"]
    ie_h = cfg["intexp"]; ii_h = cfg["intinc"]; nl_h = cfg["netloss"]

    _hist_row( 6, "Revenue",              rev[23], rev[24], rev[25], cfg.get("rev_note","Source: EDGAR"))
    _hist_row( 7, "Cost of Revenue",      cor.get(23,0), cor.get(24,0), cor[25], cfg.get("cor_note","Source / derived"))
    wa.cell(row=8, column=1, value="Gross Profit ($000s)")
    for col,ltr in [(2,"B"),(3,"C"),(4,"D")]:
        wa.cell(row=8,column=col, value=f"={ltr}6-{ltr}7").number_format = INT
    wa.cell(row=8,column=5, value="Formula: Revenue − CoR")
    wa.cell(row=9, column=1, value="Gross Margin %")
    for col,ltr in [(2,"B"),(3,"C"),(4,"D")]:
        wa.cell(row=9,column=col, value=f"={ltr}8/{ltr}6").number_format = PCT
    wa.cell(row=9,column=5, value="Formula")
    _hist_row(10, "SG&A",                sga_h[23], sga_h[24], sga_h[25], cfg.get("sga_note","Source: EDGAR"))
    _hist_row(11, "R&D",                 rd_h[23],  rd_h[24],  rd_h[25],  cfg.get("rd_note","Source: EDGAR"))
    _hist_row(12, "EBIT (Operating Loss)",ebt[23],  ebt[24],   ebt[25],   cfg.get("ebit_note","Source: EDGAR"))
    _hist_row(13, "D&A",                 da_h[23],  da_h[24],  da_h[25],  cfg.get("da_note","Source: EDGAR Cash Flow"))
    _hist_row(14, "Stock-Based Comp",    sbc_h[23], sbc_h[24], sbc_h[25], cfg.get("sbc_note","Source: EDGAR Cash Flow"))
    _hist_row(15, "CapEx (PP&E)",        cx_h[23],  cx_h[24],  cx_h[25],  cfg.get("capex_note","Source: EDGAR Cash Flow"))
    _hist_row(16, "Capitalized Software",csw_h.get(23,0), csw_h.get(24,0), csw_h[25], cfg.get("capsw_note","Source: EDGAR"))
    _hist_row(17, "Interest Expense",    ie_h[23],  ie_h[24],  ie_h[25],  cfg.get("intexp_note","Source: EDGAR"))
    _hist_row(18, "Interest Income",     ii_h[23],  ii_h[24],  ii_h[25],  cfg.get("intinc_note","Source: EDGAR"))
    _hist_row(19, "Net Loss",            nl_h[23],  nl_h[24],  nl_h[25],  cfg.get("netloss_note","Source: EDGAR"))
    if cfg.get("adj_ebitda"):
        ae = cfg["adj_ebitda"]
        _hist_row(20,"Adjusted EBITDA", None, ae.get(24), ae.get(25), cfg.get("adj_ebitda_note","Source: earnings release"))

    # ── A2. Balance Sheet ────────────────────────────────────────────────────
    _c(wa, 22, 1, f"A2.  BALANCE SHEET SNAPSHOT  (Source: {src})", bold=True)
    _hdr(wa, 23, [(1,"Line Item"),(2,"FY2024"),(3,"FY2025"),(5,"Notes")])

    bs = [
        (24,"Cash & Equivalents",         cfg["cash_24"],     cfg["cash_25"],       cfg.get("cash_note","EDGAR Cash")),
        (25,"AFS Investments (Current)",  cfg.get("afs_curr_24",0), cfg.get("afs_curr_25",0), cfg.get("afs_curr_note","EDGAR / 0 if none")),
        (26,"AFS Investments (Non-Current)",cfg.get("afs_nc_24",0), cfg.get("afs_nc_25",0),   cfg.get("afs_nc_note","EDGAR / 0 if none")),
        (27,"Total Liquidity",            "=B24+B25+B26",     "=C24+C25+C26",       "Formula: Cash + AFS"),
        (28,"Accounts Receivable",        cfg.get("ar_24",0), cfg.get("ar_25",0),   cfg.get("ar_note","EDGAR AR")),
        (29,"Goodwill",                   cfg.get("goodwill_24",0), cfg.get("goodwill_25",0), cfg.get("goodwill_note","EDGAR Goodwill")),
        (30,"Intangibles, net",           cfg.get("intg_24",0), cfg.get("intg_25",0), cfg.get("intg_note","EDGAR Intangibles")),
        (31,"Total Debt (Gross Principal)",cfg["debt_24"],    cfg["debt_25"],       cfg.get("debt_note","EDGAR Debt")),
        (32,cfg.get("post_debt_label","Debt Used in Equity Bridge"), "=B31", cfg["post_debt_25"], cfg.get("post_debt_note","")),
        (33,"Shares Outstanding",         cfg["shares_24"],   cfg["shares_25"],     cfg.get("shares_note","EDGAR diluted shares")),
        (34,cfg.get("post_shares_label","Shares Used in Bridge"), "=B33", cfg["post_shares_25"], cfg.get("post_shares_note","")),
        (35,"Federal NOL Carryforward ($K)", None,            cfg.get("nol",0),     cfg.get("nol_note","EDGAR DTA proxy")),
    ]
    for r, label, b24, c25, note in bs:
        _c(wa, r, 1, label)
        for col, val in [(2,b24),(3,c25)]:
            if val is not None:
                wa.cell(row=r, column=col, value=val).number_format = INT
        _c(wa, r, 5, note)

    # ── B. Revenue Growth ────────────────────────────────────────────────────
    _c(wa, 38, 1, "B.  REVENUE GROWTH ASSUMPTIONS  (Yellow = key input)", bold=True)
    _hdr(wa, 39, [(1,"Assumption"),(2,"Base Case"),(3,"Bull Case"),(4,"Bear Case"),(5,"Notes")])

    grow       = cfg["grow"]
    grow_notes = cfg.get("grow_notes", {})
    yr_labels  = {26:"FY2026",27:"FY2027",28:"FY2028",29:"FY2029",30:"FY2030",
                  "3135":"FY2031–2035","3640":"FY2036–2040","4145":"FY2041–2045"}
    yr_rows    = [26,27,28,29,30,"3135","3640","4145"]
    for i, yr in enumerate(yr_rows):
        row = 40 + i
        _c(wa, row, 1, f"{yr_labels[yr]} Revenue Growth Rate")
        b, u, r_ = grow[yr]
        for col, val in [(2,b),(3,u),(4,r_)]:
            _key(wa, row, col, val, PCT)
        _c(wa, row, 5, grow_notes.get(yr, ""))

    _c(wa, 49, 1, "ACTIVE SCENARIO → Enter 1 = Base Case  |  2 = Bull Case  |  3 = Bear Case")
    c50 = wa.cell(row=50, column=2, value=1)
    c50.fill = PatternFill("solid", fgColor=YELLOW)
    c50.font = Font(bold=True)
    _c(wa, 50, 1, "Active Scenario (1/2/3)")
    _c(wa, 50, 5, "Change this cell to switch scenarios on DCF sheet. 1=Base, 2=Bull, 3=Bear")

    # ── C. Margins ───────────────────────────────────────────────────────────
    _c(wa, 52, 1, "C.  MARGIN & COST ASSUMPTIONS", bold=True)
    _hdr(wa, 53, [(1,"Assumption"),(2,"Base Case"),(3,"Bull Case"),(4,"Bear Case"),(5,"Notes")])

    gm       = cfg["gm"];       gm_notes  = cfg.get("gm_notes", {})
    sga_pct  = cfg["sga_pct"];  sga_notes = cfg.get("sga_pct_notes", {})
    gm_yr_lbl = {26:"FY2026",27:"FY2027",28:"FY2028",29:"FY2029",30:"FY2030",
                 "3135":"FY2031–35","3640-45":"FY2036–45"}
    gm_keys   = [26,27,28,29,30,"3135","3640-45"]
    for i, yr in enumerate(gm_keys):
        row = 54 + i
        _c(wa, row, 1, f"Gross Margin % — {gm_yr_lbl[yr]}")
        b, u, r_ = gm[yr]
        for col, val in [(2,b),(3,u),(4,r_)]:
            _input(wa, row, col, val, PCT)
        _c(wa, row, 5, gm_notes.get(yr, ""))

    sga_yr_lbl = {26:"FY2026",27:"FY2027",28:"FY2028",29:"FY2029",30:"FY2030",
                  "3135":"FY2031–35","3640-45":"FY2036–45"}
    for i, yr in enumerate([26,27,28,29,30,"3135","3640-45"]):
        row = 61 + i
        _c(wa, row, 1, f"SG&A as % Revenue — {sga_yr_lbl[yr]}")
        b, u, r_ = sga_pct[yr]
        for col, val in [(2,b),(3,u),(4,r_)]:
            _input(wa, row, col, val, PCT)
        _c(wa, row, 5, sga_notes.get(yr, ""))

    scalar_rows = [
        (68, "R&D as % Revenue — All Years",          cfg["rd_pct"],      cfg.get("rd_pct_note","")),
        (69, "D&A as % Revenue — All Years",          cfg["da_pct"],      cfg.get("da_pct_note","")),
        (70, "CapEx as % Revenue — All Years",        cfg["capex_pct"],   cfg.get("capex_pct_note","")),
        (71, "Stock-Based Comp as % Revenue",         cfg["sbc_pct"],     cfg.get("sbc_pct_note","")),
        (72, "Change in NWC as % Rev Change",         cfg["nwc_pct"],     cfg.get("nwc_pct_note","")),
        (73, "Tax Rate (Cash)",                       cfg["cash_tax"],    cfg.get("cash_tax_note","")),
        (74, "Tax Rate (Terminal / WACC)",            cfg["term_tax"],    cfg.get("term_tax_note","US statutory rate")),
    ]
    for row, label, tup, note in scalar_rows:
        _c(wa, row, 1, label)
        b, u, r_ = tup
        for col, val in [(2,b),(3,u),(4,r_)]:
            _input(wa, row, col, val, PCT)
        _c(wa, row, 5, note)

    # ── D. WACC ──────────────────────────────────────────────────────────────
    _c(wa, 77, 1, "D.  WACC ASSUMPTIONS", bold=True)
    _hdr(wa, 78, [(1,"Assumption"),(2,"Base Case"),(3,"Bull Case"),(4,"Bear Case"),(5,"Notes")])

    wacc_rows = [
        (79,"Risk-Free Rate",      cfg["rf"],         PCT,    cfg.get("rf_note","10-yr US Treasury")),
        (80,"Equity Risk Premium", cfg["erp"],        PCT,    cfg.get("erp_note","Damodaran US ERP")),
        (81,"Beta (Levered)",      cfg["beta"],       '0.00', cfg.get("beta_note","")),
        (82,"Size / Illiquidity Premium", cfg["size_prem"], PCT, cfg.get("size_note","")),
        (83,"Cost of Equity",
            ("=B79+B81*B80+B82","=C79+C81*C80+C82","=D79+D81*D80+D82"), PCT, "Formula: Rf + Beta×ERP + Size"),
        (84,"Pre-Tax Cost of Debt", cfg["kd"],        PCT,    cfg.get("kd_note","")),
        (85,"Weight of Equity",    cfg["we"],         PCT,    cfg.get("we_note","")),
        (86,"Weight of Debt",
            ("=1-B85","=1-C85","=1-D85"),             PCT,    "Formula: 1 − Weight of Equity"),
        (87,"WACC",
            ("=B83*B85+B84*(1-D74)*B86",
             "=C83*C85+C84*(1-D74)*C86",
             "=D83*D85+D84*(1-D74)*D86"),             PCT,    "Formula: Ke×We + Kd×(1−t)×Wd"),
    ]
    for row, label, tup, fmt, note in wacc_rows:
        _c(wa, row, 1, label)
        b, u, r_ = tup
        for col, val in [(2,b),(3,u),(4,r_)]:
            cell = wa.cell(row=row, column=col, value=val)
            cell.number_format = fmt
            if row in (79,80,81,82,84,85):
                cell.fill = PatternFill("solid", fgColor=BLUE)
        _c(wa, row, 5, note)

    # ── E. Terminal Value ─────────────────────────────────────────────────────
    _c(wa, 90, 1, "E.  TERMINAL VALUE ASSUMPTIONS", bold=True)
    _hdr(wa, 91, [(1,"Assumption"),(2,"Base Case"),(3,"Bull Case"),(4,"Bear Case"),(5,"Notes")])
    tv_data = [
        (92,"Terminal Growth Rate (Gordon Growth)", cfg["tgr"],         PCT, cfg.get("tgr_note","Must be < WACC")),
        (93,"Terminal EV/EBITDA Multiple (cross-check)", cfg["ev_ebitda"], '0.0',  cfg.get("ev_ebitda_note","Comp range")),
        (94,"Terminal EBITDA Margin (for EV/EBITDA check)", cfg["ebitda_marg"], PCT, cfg.get("ebitda_marg_note","")),
    ]
    for row, label, tup, fmt, note in tv_data:
        _c(wa, row, 1, label)
        b, u, r_ = tup
        for col, val in [(2,b),(3,u),(4,r_)]:
            _key(wa, row, col, val, fmt)
        _c(wa, row, 5, note)

    # ── F. WACC Decomposition ─────────────────────────────────────────────────
    _c(wa, 97, 1, "F.  WACC DECOMPOSITION — Why WACC is high & the two levers to lower it", bold=True)
    _hdr(wa, 98, [(1,"Component"),(2,"Current Value"),(3,"Source / Formula"),
                  (4,"Contribution to WACC"),(5,"How to Lower It")])
    decomp = [
        ("Risk-Free Rate (Rf)",       "=B79","Rf × We",          "=B79*B85"),
        ("Equity Risk Premium (ERP)", "=B80","Damodaran ERP",    "=B81*B80*B85"),
        ("Beta  ← #1 LEVER",          "=B81","Levered beta",     "=B81*B80*B85"),
        ("Size/Illiquidity ← #2",     "=B82","Premium",          "=B82*B85"),
        ("Cost of Equity (Ke)",       "=B83","Rf+Beta×ERP+Size", "=B83*B85"),
        ("Cost of Debt (Kd after-tax)","=B84","Pre-tax × (1-t)", "=B84*(1-D74)*B86"),
        ("WACC",                      "=B87","Formula result",    None),
    ]
    how = cfg.get("wacc_decomp_rows", [""] * 7)
    for i, (comp, cur, src_, contrib) in enumerate(decomp):
        r = 99 + i
        _c(wa, r, 1, comp)
        _c(wa, r, 2, cur)
        _c(wa, r, 3, src_)
        if contrib:
            wa.cell(row=r, column=4, value=contrib).number_format = PCT
        _c(wa, r, 5, how[i] if i < len(how) else "")

    # ── G. WACC Compression Roadmap ───────────────────────────────────────────
    _c(wa, 107, 1, "G.  WACC COMPRESSION ROADMAP — Change Beta (row 81) and Size Premium (row 82) to model each stage", bold=True)
    _hdr(wa, 108, [(1,"Stage"),(2,"Beta"),(3,"Size Premium"),(4,"Implied WACC"),(5,"Trigger / Condition")])
    wa.cell(row=109, column=1, value="Current Base Case")
    wa.cell(row=109, column=2, value="=B81").number_format = '0.00'
    wa.cell(row=109, column=3, value="=B82").number_format = PCT
    wa.cell(row=109, column=4, value="=B87").number_format = PCT
    wa.cell(row=109, column=5, value="Pre-profit baseline")
    stages = cfg.get("wacc_stages", [])
    for i, (label, beta, size, trigger) in enumerate(stages):
        r = 110 + i
        _c(wa, r, 1, label)
        wa.cell(row=r, column=2, value=beta).number_format = '0.00'
        wa.cell(row=r, column=3, value=size).number_format = PCT
        wa.cell(row=r, column=4,
                value=f"=(B79+{beta}*B80+{size})*B85+B84*(1-D74)*B86").number_format = PCT
        _c(wa, r, 5, trigger)

    # ════════════════════════════════════════════════════════════════════════
    #  SHEET 2 — DCF Model
    # ════════════════════════════════════════════════════════════════════════
    wd = wb.create_sheet("DCF Model")
    wd.column_dimensions["A"].width = 46
    for i in range(2, 23):
        wd.column_dimensions[get_column_letter(i)].width = 12

    YEARS = ["FY2025\n(Actual)"] + [str(y) for y in range(2026, 2046)]
    _c(wd, 1, 1, f"{tk} — 20-Year Unlevered Free Cash Flow DCF Model", bold=True)
    _c(wd, 2, 1, '=IF(Assumptions!B50=1,"▶  ACTIVE: BASE CASE",'
                 'IF(Assumptions!B50=2,"▶  ACTIVE: BULL CASE","▶  ACTIVE: BEAR CASE"))')
    _c(wd, 3, 1, "Year", bold=True)
    for i, yr in enumerate(YEARS):
        c = wd.cell(row=3, column=2+i, value=yr)
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True)

    # Growth rate CHOOSE by column
    grow_asmp = {1:40,2:41,3:42,4:43,5:44,
                 6:45,7:45,8:45,9:45,10:45,
                 11:46,12:46,13:46,14:46,15:46,
                 16:47,17:47,18:47,19:47,20:47}
    gm_asmp   = {1:54,2:55,3:56,4:57,5:58,
                 6:59,7:59,8:59,9:59,10:59,
                 11:60,12:60,13:60,14:60,15:60,
                 16:60,17:60,18:60,19:60,20:60}
    sga_asmp  = {1:61,2:62,3:63,4:64,5:65,
                 6:66,7:66,8:66,9:66,10:66,
                 11:67,12:67,13:67,14:67,15:67,
                 16:67,17:67,18:67,19:67,20:67}

    def choose(ar):
        return f"=CHOOSE(Assumptions!B50,Assumptions!B{ar},Assumptions!C{ar},Assumptions!D{ar})"

    # A. Revenue
    _c(wd, 4, 1, "A.  REVENUE PROJECTION", bold=True)
    _c(wd, 5, 1, "Revenue Growth Rate (Active Scenario)")
    wd.cell(row=5, column=2, value="—")
    for i in range(1, 21):
        wd.cell(row=5, column=2+i, value=choose(grow_asmp[i])).number_format = PCT

    _c(wd, 6, 1, "Revenue ($000s)")
    wd.cell(row=6, column=2, value="=Assumptions!D6").number_format = INT
    for i in range(1, 21):
        col = 2+i; lc = get_column_letter(col); lp = get_column_letter(col-1)
        wd.cell(row=6, column=col, value=f"={lp}6*(1+{lc}5)").number_format = INT

    # B. Income Statement
    _c(wd, 7, 1, "B.  INCOME STATEMENT PROJECTION", bold=True)
    _c(wd, 8, 1, "   Gross Margin %")
    wd.cell(row=8, column=2, value="=Assumptions!D9").number_format = PCT
    for i in range(1, 21):
        wd.cell(row=8, column=2+i, value=choose(gm_asmp[i])).number_format = PCT

    _c(wd, 9, 1, "   Gross Profit ($000s)")
    wd.cell(row=9, column=2, value="=Assumptions!D8").number_format = INT
    for i in range(1, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=9, column=col, value=f"={lc}6*{lc}8").number_format = INT

    _c(wd, 10, 1, "   SG&A as % Revenue")
    wd.cell(row=10, column=2, value="=Assumptions!D10/Assumptions!D6").number_format = PCT
    for i in range(1, 21):
        wd.cell(row=10, column=2+i, value=choose(sga_asmp[i])).number_format = PCT

    _c(wd, 11, 1, "   SG&A ($000s)")
    wd.cell(row=11, column=2, value="=Assumptions!D10").number_format = INT
    for i in range(1, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=11, column=col, value=f"={lc}6*{lc}10").number_format = INT

    _c(wd, 12, 1, "   R&D as % Revenue")
    wd.cell(row=12, column=2, value="=Assumptions!D11/Assumptions!D6").number_format = PCT
    for i in range(1, 21):
        wd.cell(row=12, column=2+i, value=choose(68)).number_format = PCT

    _c(wd, 13, 1, "   R&D ($000s)")
    wd.cell(row=13, column=2, value="=Assumptions!D11").number_format = INT
    for i in range(1, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=13, column=col, value=f"={lc}6*{lc}12").number_format = INT

    _c(wd, 14, 1, "EBIT ($000s)", bold=True)
    wd.cell(row=14, column=2, value="=Assumptions!D12").number_format = INT
    for i in range(1, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=14, column=col, value=f"={lc}9-{lc}11-{lc}13").number_format = INT

    _c(wd, 15, 1, "   EBIT Margin %")
    wd.cell(row=15, column=2, value="=Assumptions!D12/Assumptions!D6").number_format = PCT
    for i in range(1, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=15, column=col, value=f"={lc}14/{lc}6").number_format = PCT

    _c(wd, 16, 1, "   Cash Tax Rate (Active Scenario)")
    wd.cell(row=16, column=2, value="=0").number_format = PCT
    for i in range(1, 21):
        wd.cell(row=16, column=2+i, value=choose(73)).number_format = PCT

    _c(wd, 17, 1, "NOPAT = EBIT × (1 − Tax Rate) ($000s)", bold=True)
    for i in range(0, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=17, column=col, value=f"={lc}14*(1-{lc}16)").number_format = INT

    # C. FCF Build
    _c(wd, 18, 1, "C.  UNLEVERED FREE CASH FLOW BUILD", bold=True)
    _c(wd, 19, 1, "   D&A as % Revenue")
    wd.cell(row=19, column=2, value="=Assumptions!D13/Assumptions!D6").number_format = PCT
    for i in range(1, 21):
        wd.cell(row=19, column=2+i, value=choose(69)).number_format = PCT

    _c(wd, 20, 1, "   (+) D&A ($000s)")
    wd.cell(row=20, column=2, value="=Assumptions!D13").number_format = INT
    for i in range(1, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=20, column=col, value=f"={lc}6*{lc}19").number_format = INT

    _c(wd, 21, 1, "   CapEx as % Revenue")
    wd.cell(row=21, column=2, value="=(Assumptions!D15+Assumptions!D16)/Assumptions!D6").number_format = PCT
    for i in range(1, 21):
        wd.cell(row=21, column=2+i, value=choose(70)).number_format = PCT

    _c(wd, 22, 1, "   (−) CapEx + Capitalized Software ($000s)")
    wd.cell(row=22, column=2, value="=Assumptions!D15+Assumptions!D16").number_format = INT
    for i in range(1, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=22, column=col, value=f"={lc}6*{lc}21").number_format = INT

    _c(wd, 23, 1, "   Δ NWC as % Revenue Change")
    wd.cell(row=23, column=2, value="=0").number_format = PCT
    for i in range(1, 21):
        wd.cell(row=23, column=2+i, value=choose(72)).number_format = PCT

    _c(wd, 24, 1, "   (−) Change in NWC ($000s)")
    wd.cell(row=24, column=2, value="=0").number_format = INT
    for i in range(1, 21):
        col=2+i; lc=get_column_letter(col); lp=get_column_letter(col-1)
        wd.cell(row=24, column=col, value=f"=({lc}6-{lp}6)*{lc}23").number_format = INT

    _c(wd, 25, 1, "UNLEVERED FCF ($000s)", bold=True)
    for i in range(0, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=25, column=col, value=f"={lc}17+{lc}20-{lc}22-{lc}24").number_format = INT

    _c(wd, 26, 1, "   UFCF Margin %")
    for i in range(0, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=26, column=col, value=f"=IF({lc}6<>0,{lc}25/{lc}6,0)").number_format = PCT

    # D. Discounting
    _c(wd, 27, 1, "D.  DISCOUNTING  (Mid-Year Convention)", bold=True)
    _c(wd, 28, 1, "WACC (Active Scenario)")
    for i in range(0, 21):
        wd.cell(row=28, column=2+i, value=choose(87)).number_format = PCT

    _c(wd, 29, 1, "   Discount Period (Mid-Year)")
    wd.cell(row=29, column=2, value="—")
    for i in range(1, 21):
        wd.cell(row=29, column=2+i, value=f"={i}-0.5").number_format = '0.0'

    _c(wd, 30, 1, "   Discount Factor")
    wd.cell(row=30, column=2, value="—")
    for i in range(1, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=30, column=col, value=f"=1/(1+{lc}28)^{lc}29").number_format = '0.0000'

    _c(wd, 31, 1, "   PV of UFCF ($000s)")
    wd.cell(row=31, column=2, value="—")
    for i in range(1, 21):
        col=2+i; lc=get_column_letter(col)
        wd.cell(row=31, column=col, value=f"={lc}25*{lc}30").number_format = INT

    _c(wd, 32, 1, "Sum of PV (UFCF) — Years 1–20 ($000s)", bold=True)
    wd.cell(row=32, column=2, value="=SUM(C31:V31)").number_format = INT

    # E. Terminal Value
    _c(wd, 33, 1, "E.  TERMINAL VALUE  (Gordon Growth Model — Primary)", bold=True)
    _c(wd, 34, 1, "   Year 20 UFCF ($000s)")
    wd.cell(row=34, column=2, value="=V25").number_format = INT
    _c(wd, 35, 1, "   Terminal Growth Rate (Active Scenario)")
    wd.cell(row=35, column=2, value=choose(92)).number_format = PCT
    _c(wd, 36, 1, "   WACC (Active Scenario)")
    wd.cell(row=36, column=2, value=choose(87)).number_format = PCT
    _c(wd, 37, 1, "Terminal Value (undiscounted) ($000s)", bold=True)
    wd.cell(row=37, column=2, value="=B34*(1+B35)/(B36-B35)").number_format = INT
    _c(wd, 38, 1, "   Terminal Value Discount Factor (Year 20)")
    wd.cell(row=38, column=2, value="=1/(1+V28)^20").number_format = '0.0000'
    _c(wd, 39, 1, "PV of Terminal Value ($000s)", bold=True)
    wd.cell(row=39, column=2, value="=B37*B38").number_format = INT
    _c(wd, 40, 1, "   Terminal Value as % of Total Enterprise Value")
    wd.cell(row=40, column=2, value="=B39/(B32+B39)").number_format = PCT

    _c(wd, 41, 1, "Terminal Value Cross-Check: EV/EBITDA Exit Multiple")
    _c(wd, 42, 1, "   Year 20 Revenue ($000s)")
    wd.cell(row=42, column=2, value="=V6").number_format = INT
    _c(wd, 43, 1, "   Terminal EBITDA Margin (Active Scenario)")
    wd.cell(row=43, column=2, value=choose(94)).number_format = PCT
    _c(wd, 44, 1, "   Implied Year 20 EBITDA ($000s)")
    wd.cell(row=44, column=2, value="=B42*B43").number_format = INT
    _c(wd, 45, 1, "   Exit EV/EBITDA Multiple (Active Scenario)")
    wd.cell(row=45, column=2, value=choose(93)).number_format = '0.0'
    _c(wd, 46, 1, "   Terminal Value — EV/EBITDA Method (undiscounted) ($000s)")
    wd.cell(row=46, column=2, value="=B44*B45").number_format = INT
    _c(wd, 47, 1, "   PV of Terminal Value — EV/EBITDA Method ($000s)")
    wd.cell(row=47, column=2, value="=B46*B38").number_format = INT

    # F. Equity Bridge
    _c(wd, 48, 1, "F.  ENTERPRISE VALUE  →  EQUITY VALUE  →  IMPLIED SHARE PRICE", bold=True)
    _c(wd, 49, 1, "Enterprise Value — Gordon Growth ($000s)", bold=True)
    wd.cell(row=49, column=2, value="=B32+B39").number_format = INT
    _c(wd, 50, 1, "   Enterprise Value — EV/EBITDA Cross-Check ($000s)")
    wd.cell(row=50, column=2, value="=B32+B47").number_format = INT
    _c(wd, 51, 1, "   (−) Net Debt (if Debt > Liquidity) ($000s)")
    wd.cell(row=51, column=2, value="=MAX(0,Assumptions!C32-Assumptions!C27)").number_format = INT
    _c(wd, 52, 1, "   (+) Net Cash (if Liquidity > Debt) ($000s)")
    wd.cell(row=52, column=2, value="=MAX(0,Assumptions!C27-Assumptions!C32)").number_format = INT
    _c(wd, 53, 1, "Equity Value — Gordon Growth ($000s)", bold=True)
    wd.cell(row=53, column=2, value="=B49-B51+B52").number_format = INT
    _c(wd, 54, 1, "   Diluted Shares Outstanding")
    wd.cell(row=54, column=2, value="=Assumptions!C34").number_format = INT
    _c(wd, 55, 1, "IMPLIED SHARE PRICE ($)", bold=True)
    wd.cell(row=55, column=2, value="=IF(B54>0,(B53*1000)/B54,0)").number_format = DOL

    # G. Scenario Summary
    _c(wd, 57, 1, "G.  SCENARIO SUMMARY  (Run each scenario by changing Active Scenario on Assumptions sheet)", bold=True)
    _hdr(wd, 58, [(1,"Metric"),(2,"Base Case\n(Scenario 1)"),(3,"Bull Case\n(Scenario 2)"),(4,"Bear Case\n(Scenario 3)")])
    for col in [2,3,4]:
        wd.cell(row=58, column=col).alignment = Alignment(wrap_text=True)

    scen = [
        (59,"FY2026E Revenue ($000s)",      "=B102","=C102","=D102",INT),
        (60,"FY2026E Revenue Growth %",     "=Assumptions!B40","=Assumptions!C40","=Assumptions!D40",PCT),
        (61,"FY2030E Revenue ($000s)",      "=B106","=C106","=D106",INT),
        (62,"FY2045E Revenue ($000s)",      "=B121","=C121","=D121",INT),
        (63,"FY2026E Gross Margin %",       "=Assumptions!B54","=Assumptions!C54","=Assumptions!D54",PCT),
        (64,"FY2026E EBIT ($000s)",
            "=B102*(Assumptions!B54-Assumptions!B61-Assumptions!B68)",
            "=C102*(Assumptions!C54-Assumptions!C61-Assumptions!C68)",
            "=D102*(Assumptions!D54-Assumptions!D61-Assumptions!D68)",INT),
        (65,"WACC",                         "=Assumptions!B87","=Assumptions!C87","=Assumptions!D87",PCT),
        (66,"Terminal Growth Rate",         "=Assumptions!B92","=Assumptions!C92","=Assumptions!D92",PCT),
        (67,"Sum PV of FCFs ($000s)",       "=B162","=C162","=D162",INT),
        (68,"PV of Terminal Value ($000s)", "=B164","=C164","=D164",INT),
        (69,"Enterprise Value ($000s)",     "=B165","=C165","=D165",INT),
        (70,"Equity Value ($000s)",         "=B166","=C166","=D166",INT),
        (71,"Implied Share Price ($)",      "=B167","=C167","=D167",DOL),
    ]
    for r, label, b, c_, d, fmt in scen:
        _c(wd, r, 1, label)
        for col, val in [(2,b),(3,c_),(4,d)]:
            wd.cell(row=r, column=col, value=val).number_format = fmt

    # H. Sensitivity
    _c(wd, 74, 1, "H.  SENSITIVITY: Implied Share Price ($)  —  WACC vs Terminal Growth Rate  (Base Case)", bold=True)
    _hdr(wd, 75, [(1,"WACC \\ TGR →"),(2,"2.0%"),(3,"2.5%"),(4,"3.0%"),
                   (5,"3.5%"),(6,"4.0%"),(7,"4.5%")])
    for ri, w in enumerate([0.10,0.12,0.14,0.16,0.18,0.20]):
        row = 76+ri
        wd.cell(row=row, column=1, value=f"{int(w*100)}%").font = Font(bold=True)
        for ci, g in enumerate([0.020,0.025,0.030,0.035,0.040,0.045]):
            wd.cell(row=row, column=2+ci, value=_sens(w,g)).number_format = DOL

    # Scenario Engine (rows 100-167)
    _c(wd, 100, 1, "── SCENARIO ENGINE (helper rows – do not edit) ──", bold=True)
    _hdr(wd, 101, [(1,"Metric"),(2,"Base"),(3,"Bull"),(4,"Bear")])

    ga_map = [40,41,42,43,44,45,45,45,45,45,46,46,46,46,46,47,47,47,47,47]
    for i, ar in enumerate(ga_map):
        row = 102+i; yr = 2026+i
        _c(wd, row, 1, f"Rev {yr}")
        for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
            if i == 0:
                v = f"=Assumptions!D6*(1+Assumptions!{sc}{ar})"
            else:
                v = f"={sc}{101+i}*(1+Assumptions!{sc}{ar})"
            wd.cell(row=row, column=col, value=v).number_format = INT

    gm_map  = [54,55,56,57,58,59,59,59,59,59,60,60,60,60,60,60,60,60,60,60]
    sga_map = [61,62,63,64,65,66,66,66,66,66,67,67,67,67,67,67,67,67,67,67]
    for i in range(20):
        row = 122+i; yr = 2026+i
        _c(wd, row, 1, f"UFCF {yr}")
        for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
            rv = f"{sc}{102+i}"
            pr = "Assumptions!D6" if i==0 else f"{sc}{101+i}"
            v = (f"=(({rv}*Assumptions!{sc}{gm_map[i]}-{rv}*Assumptions!{sc}{sga_map[i]}"
                 f"-{rv}*Assumptions!{sc}68)*(1-Assumptions!{sc}73))"
                 f"+({rv}*Assumptions!{sc}69)"
                 f"-({rv}*Assumptions!{sc}70)"
                 f"-(({rv}-{pr})*Assumptions!{sc}72)")
            wd.cell(row=row, column=col, value=v).number_format = INT

    for i in range(20):
        row = 142+i; yr = 2026+i
        _c(wd, row, 1, f"PV_UFCF {yr}")
        for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
            wd.cell(row=row, column=col,
                    value=f"={sc}{122+i}/(1+Assumptions!{sc}87)^{i+0.5}").number_format = INT

    _c(wd, 162, 1, "Sum PV_UFCF", bold=True)
    for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
        wd.cell(row=162, column=col, value=f"=SUM({sc}142:{sc}161)").number_format = INT

    _c(wd, 164, 1, "PV of Terminal Value", bold=True)
    for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
        wd.cell(row=164, column=col,
                value=f"={sc}141*(1+Assumptions!{sc}92)/(Assumptions!{sc}87-Assumptions!{sc}92)/(1+Assumptions!{sc}87)^20"
               ).number_format = INT

    _c(wd, 165, 1, "Enterprise Value", bold=True)
    for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
        wd.cell(row=165, column=col, value=f"={sc}162+{sc}164").number_format = INT

    nd = "MAX(0,Assumptions!C32-Assumptions!C27)"
    nc = "MAX(0,Assumptions!C27-Assumptions!C32)"
    _c(wd, 166, 1, "Equity Value", bold=True)
    for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
        wd.cell(row=166, column=col, value=f"={sc}165-{nd}+{nc}").number_format = INT

    _c(wd, 167, 1, "Implied Share Price ($)", bold=True)
    for col, sc in [(2,"B"),(3,"C"),(4,"D")]:
        wd.cell(row=167, column=col,
                value=f"=IF(Assumptions!C34>0,({sc}166*1000)/Assumptions!C34,0)").number_format = DOL

    return wb


# ── Ticker configs ────────────────────────────────────────────────────────────

def cfg_tem() -> dict:
    return dict(
        ticker="TEM", cik="0001717115",
        source_note="TEM 10-K FY2025, SEC EDGAR CIK 0001717115",
        rev={23:531822, 24:693398, 25:1271789},
        cor={23:340802, 24:479804, 25:619999},
        sga={23:296760, 24:755351, 25:731738},
        rd ={23:90343,  24:149325, 25:172924},
        ebit={23:-196083,24:-691082,25:-252872},
        da ={23:21279,  24:26356,  25:32054},
        sbc={23:0,      24:534138, 25:124747},
        capex={23:34608,24:22121,  25:21049},
        capsw={23:6137, 24:4292,   25:5535},
        intexp={23:46869,24:53653, 25:70267},
        intinc={23:7601, 24:11084, 25:12628},
        netloss={23:-214118,24:-705809,25:-245028},
        cash_24=340954, cash_25=604787,
        afs_curr_24=0,  afs_curr_25=0,
        afs_nc_24=0,    afs_nc_25=0,
        ar_24=154819,   ar_25=311170,
        goodwill_24=73343,  goodwill_25=470211,
        intg_24=11716,      intg_25=355253,
        debt_24=168192,     debt_25=208672,
        post_debt_25=208672,
        post_debt_label="Debt Used in Bridge (Conv. Notes)",
        post_debt_note="Convertible notes; no near-term forced conversion",
        shares_24=119849000, shares_25=174264000,
        post_shares_25=174264000,
        post_shares_label="Shares Used in Bridge",
        post_shares_note="FY2025 diluted weighted-average × 1000",
        nol=727301,
        nol_note="EDGAR DeferredTaxAssetsGross FY2025",
        grow={26:(0.35,0.45,0.20),27:(0.30,0.38,0.15),28:(0.25,0.32,0.12),
              29:(0.22,0.28,0.10),30:(0.20,0.25,0.08),
              "3135":(0.17,0.22,0.08),"3640":(0.11,0.15,0.06),"4145":(0.07,0.10,0.04)},
        grow_notes={26:"Moderation from 83% FY2025; Ambry now integrated",
                    27:"Data platform + genomics ramp",30:"TAM penetration"},
        gm={26:(0.53,0.57,0.48),27:(0.56,0.60,0.50),28:(0.59,0.63,0.52),
            29:(0.62,0.66,0.54),30:(0.64,0.69,0.56),
            "3135":(0.67,0.72,0.58),"3640-45":(0.70,0.76,0.60)},
        gm_notes={26:"FY2025=51.2%; Ambry + AI platform scale"},
        sga_pct={26:(0.48,0.42,0.56),27:(0.42,0.36,0.52),28:(0.37,0.31,0.47),
                 29:(0.33,0.27,0.43),30:(0.28,0.22,0.40),
                 "3135":(0.22,0.17,0.34),"3640-45":(0.18,0.13,0.28)},
        sga_pct_notes={26:"FY2025=57.5% incl SBC; heavy sales investment"},
        rd_pct=(0.11,0.13,0.10), rd_pct_note="FY2025=13.6%; normalizing",
        da_pct=(0.08,0.07,0.09), da_pct_note="Elevated post-Ambry intangible amort",
        capex_pct=(0.04,0.03,0.05), capex_pct_note="Asset-light; FY2025=2.1% rev",
        sbc_pct=(0.10,0.09,0.12),
        nwc_pct=(0.06,0.04,0.09), nwc_pct_note="AR-heavy; genomics payer lag",
        cash_tax=(0.00,0.00,0.02), cash_tax_note="$727M DTA / large NOL",
        term_tax=(0.21,0.21,0.21),
        rf=(0.043,0.040,0.046), erp=(0.055,0.050,0.060),
        beta=(1.70,1.40,2.10),  size_prem=(0.020,0.010,0.030),
        kd=(0.055,0.050,0.065), we=(0.85,0.90,0.80),
        rf_note="10-yr US Treasury May 2026",
        beta_note="Health-AI growth co.; post-IPO volatility",
        tgr=(0.035,0.040,0.025), ev_ebitda=(20,28,12), ebitda_marg=(0.28,0.35,0.18),
        tgr_note="Must be < WACC; healthcare AI long-run GDP+ growth",
        wacc_stages=[
            ("Stage 1 — FCF Break-Even",1.30,0.015,"TEM FCF-positive ~FY2029E"),
            ("Stage 2 — Proven Platform",1.00,0.005,"ARR visible; genomics recurring revenue"),
            ("Stage 3 — Re-Rating ($50+)",0.80,0.000,"Debt-free, profitable 3+ yrs"),
            ("Blue Sky — Epic/PLTR Comp",0.60,0.000,"Tempus moat = health-AI infrastructure"),
        ],
    )


def cfg_bbai() -> dict:
    return dict(
        ticker="BBAI", cik="0001836981",
        source_note="BBAI 10-K FY2025, filed 03/02/2026",
        rev={23:155164, 24:158236, 25:127672},
        cor={23:114563, 24:113016, 25:99194},
        sga={23:71057,  24:80040,  25:95132},
        rd ={23:5035,   24:10863,  25:16752},
        ebit={23:-39034,24:-133420,25:-213897},
        da ={23:7901,   24:11873,  25:15281},
        sbc={23:18671,  24:21127,  25:23330},
        capex={23:2,    24:484,    25:525},
        capsw={23:3828, 24:10630,  25:3841},
        intexp={23:24877,24:25647, 25:18116},
        intinc={23:392,  24:2293,  25:13253},
        netloss={23:-70657,24:-295547,25:-293914},
        adj_ebitda={24:-2397,25:-35140},
        adj_ebitda_note="Source: BBAI Q4 2025 Press Release",
        cash_24=50141,  cash_25=87126,
        afs_curr_24=0,  afs_curr_25=200461,
        afs_nc_24=0,    afs_nc_25=173949,
        ar_24=38953,    ar_25=22703,
        goodwill_24=119081, goodwill_25=241100,
        intg_24=119119,     intg_25=139470,
        debt_24=200818,     debt_25=142273,
        post_debt_25="=C31-124605",
        post_debt_label="Post-Jan-2026 Debt (2026 Notes Only)",
        post_debt_note="Formula: 2029 Notes converted Jan 2026; only 2026 Notes ($17.7M) remain",
        shares_24=251554378, shares_25=436955655,
        post_shares_25="=C33+38100000",
        post_shares_label="Post-Jan-2026 Shares (est.)",
        post_shares_note="Formula: FY25 shares + ~38.1M from 2029 Note conversions",
        nol=339600,
        nol_note="Source: BBAI 10-K FY2025, p.125 (unlimited carryforward)",
        grow={26:(0.17,0.25,0.05),27:(0.20,0.30,0.10),28:(0.22,0.30,0.12),
              29:(0.25,0.30,0.15),30:(0.22,0.28,0.12),
              "3135":(0.18,0.22,0.10),"3640":(0.12,0.15,0.07),"4145":(0.08,0.10,0.05)},
        grow_notes={26:"Mgmt 2026 guide: $135M–$165M (~17% mid)",
                    27:"Ask Sage ($229M acquisition) ramp + Middle East expansion"},
        gm={26:(0.24,0.27,0.21),27:(0.27,0.31,0.23),28:(0.30,0.34,0.25),
            29:(0.33,0.37,0.27),30:(0.35,0.40,0.29),
            "3135":(0.38,0.44,0.31),"3640-45":(0.42,0.50,0.33)},
        gm_notes={26:"Ask Sage SaaS mix improves blended margin; 2025=22.3%"},
        sga_pct={26:(0.65,0.58,0.72),27:(0.55,0.48,0.65),28:(0.48,0.40,0.58),
                 29:(0.42,0.34,0.52),30:(0.38,0.30,0.47),
                 "3135":(0.30,0.24,0.40),"3640-45":(0.25,0.20,0.35)},
        sga_pct_notes={26:"FY2025=75%; assumes sales investment continues"},
        rd_pct=(0.10,0.12,0.08), rd_pct_note="FY2025=13%; normalizing",
        da_pct=(0.10,0.09,0.11), da_pct_note="FY2025=12%; intangible amort declines",
        capex_pct=(0.03,0.02,0.04), capex_pct_note="Asset-light; FY2025 CapEx+CapSW=3.4%",
        sbc_pct=(0.10,0.09,0.12),
        nwc_pct=(0.05,0.03,0.08), nwc_pct_note="Low NWC intensity for government services",
        cash_tax=(0.00,0.00,0.02), cash_tax_note="$339.6M federal NOL; cash taxes ~0%",
        term_tax=(0.21,0.21,0.21),
        rf=(0.043,0.040,0.046), erp=(0.055,0.050,0.060),
        beta=(2.20,1.90,2.50),  size_prem=(0.030,0.020,0.040),
        kd=(0.060,0.055,0.070), we=(0.85,0.90,0.80),
        rf_note="10-yr US Treasury April 2026",
        beta_note="High-beta small-cap defense AI; BBAI goodwill impairment used 12% rate",
        tgr=(0.03,0.035,0.025), ev_ebitda=(15,20,10), ebitda_marg=(0.20,0.26,0.14),
        tgr_note="Must be < WACC; long-run nominal GDP proxy",
        wacc_stages=[
            ("Stage 1 — FCF Break-Even",1.60,0.020,"BBAI hits FCF-positive in 2028; ATM ends"),
            ("Stage 2 — Proven Platform",1.30,0.010,"ARR growth visible; Ask Sage recurring"),
            ("Stage 3 — Re-Rating ($10)",1.00,0.000,"Debt-free, profitable 3+ yrs"),
            ("Blue Sky — Palantir Comp",0.85,0.000,"BBAI beta converges to Palantir peers"),
        ],
    )


def cfg_rgti() -> dict:
    """Rigetti Computing (RGTI) — CIK 0001838359, FY2025 10-K."""
    # Revenue is DECLINING ($13M→$7M) — quantum pre-commercialisation phase
    # Expenses are largely fixed, dwarfing revenue; huge cash burn
    return dict(
        ticker="RGTI", cik="0001838359",
        source_note="RGTI 10-K FY2025, SEC EDGAR CIK 0001838359",
        rev_note   ="EDGAR RevenueFromContractWithCustomerIncludingAssessedTax",
        cor_note   ="EDGAR CostOfRevenue",
        sga_note   ="EDGAR GeneralAndAdministrativeExpense",
        rd_note    ="EDGAR ResearchAndDevelopmentExpense",
        ebit_note  ="EDGAR OperatingIncomeLoss",
        da_note    ="EDGAR DepreciationDepletionAndAmortization (cash-flow stmt)",
        sbc_note   ="EDGAR AllocatedShareBasedCompensationExpense",
        capex_note ="EDGAR PaymentsToAcquirePropertyPlantAndEquipment",
        intexp_note="EDGAR InterestExpense (0 in FY2024/25 — debt fully retired)",
        intinc_note="Not separately tagged in XBRL",
        netloss_note="EDGAR NetIncomeLoss",
        # ── Historical income statement ($K) ──────────────────────────────
        rev    ={23:12008,  24:10790,  25:7088},
        cor    ={23:2800,   24:5093,   25:5024},
        sga    ={23:27744,  24:24457,  25:25379},
        rd     ={23:52768,  24:49750,  25:61345},
        ebit   ={23:-72295, 24:-68510, 25:-84660},
        da     ={23:7400,   24:6900,   25:8100},
        sbc    ={23:12409,  24:13069,  25:17605},
        capex  ={23:9059,   24:11098,  25:18676},
        capsw  ={23:0,      24:0,      25:0},
        intexp ={23:5779,   24:0,      25:0},
        intinc ={23:0,      24:0,      25:0},
        netloss={23:-75107, 24:-200988,25:-216210},
        # ── Balance sheet ($K) ────────────────────────────────────────────
        cash_24=67674,   cash_25=44851,
        afs_curr_24=0,   afs_curr_25=0,
        afs_nc_24=0,     afs_nc_25=0,
        ar_24=2427,      ar_25=2551,
        goodwill_24=0,   goodwill_25=0,
        intg_24=0,       intg_25=0,
        debt_24=0,       debt_25=0,      # fully retired term loan
        post_debt_25=0,
        post_debt_label="Post-FY2025 Debt",
        post_debt_note ="Debt-free as of FY2024; term loan fully retired",
        shares_24=184666000, shares_25=309763000,
        post_shares_25=309763000,
        post_shares_label="Shares Used in Bridge",
        post_shares_note ="FY2025 diluted wtd-avg (significant ongoing dilution expected)",
        nol=131826,
        nol_note="EDGAR DeferredTaxAssetsGross FY2025 ($K proxy for NOL)",
        # ── Revenue growth (Base / Bull / Bear) ───────────────────────────
        # Revenue was $7M in FY2025; bull case = quantum commercial adoption
        grow={
            26:(0.50, 1.00,-0.10),   # base +50%, bull +100%, bear -10%
            27:(0.80, 2.00, 0.10),
            28:(1.20, 2.50, 0.20),
            29:(1.00, 2.00, 0.30),
            30:(0.80, 1.50, 0.25),
            "3135":(0.60, 1.00, 0.18),
            "3640":(0.35, 0.60, 0.10),
            "4145":(0.15, 0.25, 0.07),
        },
        grow_notes={
            26:"Revenue declining in FY2025; stabilisation + first commercial ramp",
            27:"Quantum advantage demonstrations driving enterprise pilots",
            28:"Multi-qubit systems reaching commercial-grade fidelity",
            29:"Hyperscaler + pharma / finance contract wins",
            30:"TAM penetration; Rigetti platform-as-a-service scaling",
            "3135":"Quantum-classical hybrid workloads mainstream",
        },
        # ── Gross margin (low now; recovers with software/service mix) ────
        gm={
            26:(0.32,0.42,0.20), 27:(0.45,0.60,0.25), 28:(0.55,0.68,0.30),
            29:(0.62,0.72,0.35), 30:(0.65,0.74,0.40),
            "3135":(0.68,0.76,0.45), "3640-45":(0.70,0.78,0.50),
        },
        gm_notes={26:"FY2025=29%; hardware-heavy mix; software shift begins"},
        # ── SG&A % — absolute SGA ~$25M fixed; % falls as rev scales ─────
        sga_pct={
            26:(2.20,1.40,3.10), 27:(1.10,0.65,2.40), 28:(0.58,0.38,1.40),
            29:(0.35,0.22,0.85), 30:(0.24,0.16,0.55),
            "3135":(0.17,0.12,0.35), "3640-45":(0.12,0.08,0.26),
        },
        sga_pct_notes={26:"FY2025 SGA/Rev=358%; absolute SGA ~$25M semi-fixed"},
        # ── Other cost ratios ─────────────────────────────────────────────
        rd_pct    =(0.45,0.35,0.55), rd_pct_note="FY2025 R&D/Rev=866%; declines as rev scales",
        da_pct    =(0.45,0.38,0.52), da_pct_note="FY2025 DA/Rev=114%; reflects quantum hardware depreciation",
        capex_pct =(0.15,0.10,0.20), capex_pct_note="FY2025=263% of tiny revenue; quantum system capex",
        sbc_pct   =(0.12,0.10,0.15), sbc_pct_note="FY2025 SBC=$17.6M; normalising as rev grows",
        nwc_pct   =(0.03,0.02,0.05),
        cash_tax  =(0.00,0.00,0.00), cash_tax_note="$132M DTA; NOL shields all near-term income",
        term_tax  =(0.21,0.21,0.21),
        # ── WACC (high: pre-revenue, binary outcome) ──────────────────────
        rf=(0.043,0.040,0.046), erp=(0.055,0.050,0.060),
        beta=(2.50,2.00,3.00), size_prem=(0.035,0.020,0.050),
        kd=(0.050,0.045,0.065), we=(0.90,0.95,0.85),
        rf_note   ="10-yr US Treasury May 2026",
        beta_note ="Deep-tech pre-revenue; quantum binary outcome; ~2.5–3.0 beta",
        kd_note   ="No current debt; modelled for future capital raises",
        we_note   ="Equity-funded; may issue convertibles for quantum hardware buildout",
        # ── Terminal value ────────────────────────────────────────────────
        tgr=(0.035,0.045,0.020),
        ev_ebitda=(25,40,8),
        ebitda_marg=(0.20,0.35,0.05),
        tgr_note="Quantum-AI platform; long-run GDP+ growth if commercialised",
        ev_ebitda_note="Quantum platform comps: 20–50× if winner; 5–10× if niche",
        ebitda_marg_note="Requires platform shift; highly binary",
        # ── WACC compression roadmap ──────────────────────────────────────
        wacc_stages=[
            ("Stage 1 — First Revenue Scale ($50M)",1.80,0.025,"Recurring quantum-cloud ARR materialises"),
            ("Stage 2 — Break-Even Proof",          1.40,0.015,"FCF-positive; quantum-error-correction milestone"),
            ("Stage 3 — Hyperscaler Partnerships",  1.10,0.005,"AWS/Azure/GCP embedded quantum layer"),
            ("Blue Sky — Quantum Category Leader",  0.80,0.000,"De-facto standard for quantum computing infra"),
        ],
    )


# ── CLI entry ─────────────────────────────────────────────────────────────────

CONFIGS = {"TEM": cfg_tem, "BBAI": cfg_bbai, "RGTI": cfg_rgti}

def build(ticker: str, out_path: Path = None) -> Path:
    ticker = ticker.upper()
    if ticker in CONFIGS:
        cfg = CONFIGS[ticker]()
    else:
        raise ValueError(f"Unknown ticker: {ticker}")
    if out_path is None:
        out_path = Path(__file__).parent / f"{ticker}_dcf.xlsx"
    wb = build_workbook(cfg)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--ticker", required=True, choices=["TEM","BBAI","RGTI"],
                        help="Ticker to build DCF for")
    args = parser.parse_args()
    path = build(args.ticker)
    print(f"Saved: {path}")
